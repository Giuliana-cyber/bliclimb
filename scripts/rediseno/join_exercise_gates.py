#!/usr/bin/env python3
"""
JOIN determinístico Exercise → Gates via APP_Relationships.

Fase 1 · Task F1.3 · aprobado por Giuliana el 2026-07-15.

Fuente:
  - APP_ExerciseCatalog (582 filas, col `gates` completamente vacía)
  - APP_Relationships (1970 filas). Relations relevantes:
      * BLOCKED_BY   (402 filas)  — ejercicio bloqueado por gate
      * CONTROLLED_BY (36 filas)  — ejercicio controlado por gate (adapta)
      * GOVERNED_BY  (19 filas)   — ejercicio regido por gate a nivel política

Salida:
  - reporte JSON con 582 asignaciones ejercicio→[gate_ids]
  - reporte CSV amigable para revisión de Giuliana
  - reporte MD con estadísticas (cobertura, orphan gates, dispersión)

Regla:
  - Fail-closed: si un `to_id` en Relationships no aparece en APP_GateCatalog,
    se registra como issue pero NO se emite en el mapping.
  - Deterministic: mismo input → mismo output (sort estable por ID).
  - No se modifica el v3.0 ni el v3.1. Solo se emiten reportes.

Uso:
  python scripts/rediseno/join_exercise_gates.py
    [--source docs/biblioteca-maestra-v3.1.xlsx]
    [--outdir docs/rediseno/reports/f13_join_gates]

El reporte queda en docs/rediseno/reports/ para revisión editorial de Giuliana
antes de que el mapping se persista en el catálogo.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any


# Relaciones canónicas que conectan Exercise → Gate como GATES REALES.
# Un gate de seguridad debe ser deliberado — evitamos seguridad falsa por
# vínculo casual. Aprobado por Giuliana 2026-07-15.
GATE_LINK_RELATIONS = {"BLOCKED_BY", "CONTROLLED_BY", "GOVERNED_BY"}

# Relación DÉBIL: RELATED_TO conecta EX→GT 216 veces adicionales, pero
# es semánticamente floja ("relacionado con"). Se persiste en columna
# APARTE `gates_related_to` para uso en búsqueda/chat, NUNCA como gate real.
# Cuando Giuliana cure una categoría, promueve los relevantes a canónicas.
# Aprobado por Giuliana 2026-07-15 (Opción C).
WEAK_GATE_LINK_RELATIONS = {"RELATED_TO"}


def load_workbook_readonly(path: Path):
    """Carga xlsx en modo read_only para bypasear el bug RoadmapStatusTable."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl no está instalado. pip install openpyxl")
    return load_workbook(path, data_only=True, read_only=True)


def sheet_to_dicts(ws) -> list[dict[str, str]]:
    """Convierte una hoja a lista de dicts usando la primera fila como headers."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return []
    headers = [h if h is not None else "" for h in headers]
    out: list[dict[str, str]] = []
    for r in rows_iter:
        if not any(c is not None for c in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            if i < len(r) and r[i] is not None:
                row[h] = str(r[i])
            else:
                row[h] = ""
        out.append(row)
    return out


def build_join(
    exercises: list[dict[str, str]],
    gates: list[dict[str, str]],
    relationships: list[dict[str, str]],
) -> dict[str, Any]:
    """
    Construye el mapping ejercicio→[gate_ids] aplicando las 3 relaciones
    canónicas. Devuelve dict con `mapping`, `stats`, `issues`.
    """
    exercise_ids = {ex["exercise_id"] for ex in exercises if ex.get("exercise_id")}
    gate_ids = {g["gate_id"] for g in gates if g.get("gate_id")}

    # from_id → mapping por relation. Las 3 canónicas + la débil (columna
    # separada, ver docstring en constantes arriba).
    mapping: dict[str, dict[str, list[str]]] = defaultdict(
        lambda: {
            "BLOCKED_BY": [],
            "CONTROLLED_BY": [],
            "GOVERNED_BY": [],
            "RELATED_TO": [],  # débil · columna aparte en el export
        }
    )
    issues: list[dict[str, str]] = []

    used_relations = Counter()
    all_gate_link_relations = GATE_LINK_RELATIONS | WEAK_GATE_LINK_RELATIONS

    for rel in relationships:
        rtype = rel.get("relation", "").strip()
        if rtype not in all_gate_link_relations:
            continue
        to_id_raw = rel.get("to_id", "").strip()
        # WEAK_GATE_LINK_RELATIONS incluye RELATED_TO que se usa entre muchos
        # tipos de entidad; solo lo consideramos cuando el to_id es un gate.
        if rtype in WEAK_GATE_LINK_RELATIONS and not to_id_raw.startswith("GT-"):
            continue
        used_relations[rtype] += 1
        from_id = rel.get("from_id", "").strip()
        to_id = rel.get("to_id", "").strip()

        # from_id debe estar en Exercises (es exercise→gate)
        if from_id not in exercise_ids:
            issues.append(
                {
                    "kind": "from_id_not_exercise",
                    "relationship_id": rel.get("relationship_id", ""),
                    "from_id": from_id,
                    "relation": rtype,
                    "to_id": to_id,
                }
            )
            continue

        # to_id debe estar en Gates
        if to_id not in gate_ids:
            issues.append(
                {
                    "kind": "to_id_not_gate",
                    "relationship_id": rel.get("relationship_id", ""),
                    "from_id": from_id,
                    "relation": rtype,
                    "to_id": to_id,
                }
            )
            continue

        # Idempotente: no duplicar el mismo (from_id, relation, to_id)
        bucket = mapping[from_id][rtype]
        if to_id not in bucket:
            bucket.append(to_id)

    # Ordenar deterministicamente cada lista de gate_ids
    for from_id in mapping:
        for rtype in mapping[from_id]:
            mapping[from_id][rtype].sort()

    # Estadísticas — cobertura por gates CANÓNICOS (las 3 duras). RELATED_TO
    # es información secundaria, no cuenta como cobertura de seguridad.
    coverage = {
        "exercises_total": len(exercise_ids),
        "exercises_with_at_least_one_canonical_gate": sum(
            1
            for ex_id in exercise_ids
            if any(mapping.get(ex_id, {}).get(r) for r in GATE_LINK_RELATIONS)
        ),
        "exercises_with_zero_canonical_gates": 0,
        "exercises_with_only_weak_gates": 0,
        "exercises_with_at_least_one_weak_gate": sum(
            1
            for ex_id in exercise_ids
            if any(mapping.get(ex_id, {}).get(r) for r in WEAK_GATE_LINK_RELATIONS)
        ),
        "gates_total": len(gate_ids),
        "gates_referenced_by_canonical": 0,
        "gates_referenced_only_by_weak": 0,
        "orphan_gates_never_referenced": 0,
        "relations_consumed": dict(used_relations),
        "issues_total": len(issues),
    }
    coverage["exercises_with_zero_canonical_gates"] = (
        coverage["exercises_total"]
        - coverage["exercises_with_at_least_one_canonical_gate"]
    )
    coverage["exercises_with_only_weak_gates"] = sum(
        1
        for ex_id in exercise_ids
        if not any(mapping.get(ex_id, {}).get(r) for r in GATE_LINK_RELATIONS)
        and any(mapping.get(ex_id, {}).get(r) for r in WEAK_GATE_LINK_RELATIONS)
    )

    canonical_gates = {
        gid
        for ex_id in mapping
        for r in GATE_LINK_RELATIONS
        for gid in mapping[ex_id][r]
    }
    weak_gates = {
        gid
        for ex_id in mapping
        for r in WEAK_GATE_LINK_RELATIONS
        for gid in mapping[ex_id][r]
    }
    all_referenced_gates = canonical_gates | weak_gates
    coverage["gates_referenced_by_canonical"] = len(canonical_gates)
    coverage["gates_referenced_only_by_weak"] = len(weak_gates - canonical_gates)
    coverage["orphan_gates_never_referenced"] = len(gate_ids - all_referenced_gates)

    # Distribución de gates canónicos por ejercicio
    dist_total_gates: Counter = Counter()
    for ex_id in exercise_ids:
        n = sum(len(mapping.get(ex_id, {}).get(r, [])) for r in GATE_LINK_RELATIONS)
        dist_total_gates[n] += 1
    coverage["distribution_canonical_gates_per_exercise"] = dict(sorted(dist_total_gates.items()))

    return {
        "mapping": {ex_id: dict(v) for ex_id, v in mapping.items()},
        "stats": coverage,
        "issues": issues,
        "orphan_gates": sorted(gate_ids - all_referenced_gates),
    }


def write_reports(join: dict[str, Any], exercises: list[dict[str, str]], outdir: Path) -> None:
    outdir.mkdir(parents=True, exist_ok=True)

    # 1. JSON con el mapping completo (source of truth para revisión programática)
    with (outdir / "exercise_gates_mapping.json").open("w") as f:
        json.dump(join, f, indent=2, ensure_ascii=False, sort_keys=True)

    # 2. CSV amigable para revisión editorial de Giuliana
    ex_by_id = {ex["exercise_id"]: ex for ex in exercises if ex.get("exercise_id")}
    with (outdir / "exercise_gates_review.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "exercise_id",
                "exercise_name",
                "exercise_category",
                "risk_level",
                "gates_blocked_by",
                "gates_controlled_by",
                "gates_governed_by",
                "gates_canonical_count",
                "gates_related_to",  # débil · columna aparte
                "gates_weak_count",
            ]
        )
        # Sort por exercise_id para determinismo
        for ex_id in sorted(ex_by_id.keys()):
            ex = ex_by_id[ex_id]
            m = join["mapping"].get(ex_id, {})
            b = m.get("BLOCKED_BY", [])
            c = m.get("CONTROLLED_BY", [])
            g = m.get("GOVERNED_BY", [])
            r = m.get("RELATED_TO", [])
            w.writerow(
                [
                    ex_id,
                    ex.get("name_es", ""),
                    ex.get("category", ""),
                    ex.get("risk_level", ""),
                    "; ".join(b),
                    "; ".join(c),
                    "; ".join(g),
                    len(b) + len(c) + len(g),
                    "; ".join(r),
                    len(r),
                ]
            )

    # 3. Issues CSV separado
    with (outdir / "issues.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["kind", "relationship_id", "from_id", "relation", "to_id"])
        for issue in join["issues"]:
            w.writerow([
                issue.get("kind", ""),
                issue.get("relationship_id", ""),
                issue.get("from_id", ""),
                issue.get("relation", ""),
                issue.get("to_id", ""),
            ])

    # 4. Reporte MD ejecutivo
    stats = join["stats"]
    orph = join["orphan_gates"]
    md_lines = [
        "# Reporte JOIN Exercise → Gates (v2 · con distinción canónico/débil)",
        "",
        f"**Fuente**: `docs/biblioteca-maestra-LATEST.xlsx`",
        f"**Fecha**: 2026-07-15 · Fase 1 · Task F1.3",
        f"**Relaciones canónicas** (gate real de seguridad): {sorted(GATE_LINK_RELATIONS)}",
        f"**Relación débil** (columna aparte, uso en chat/búsqueda): {sorted(WEAK_GATE_LINK_RELATIONS)}",
        "",
        "> Un gate de seguridad debe ser deliberado, no un vínculo casual — evitamos seguridad falsa.",
        "> RELATED_TO se persiste en `gates_related_to` para uso en búsqueda/chat, NUNCA como gate real.",
        "> Cuando Giuliana cure cada categoría, promueve los relevantes a canónicas.",
        "",
        "## Cobertura por gates CANÓNICOS (los que gatean seguridad real)",
        "",
        f"- Ejercicios totales: **{stats['exercises_total']}**",
        f"- Ejercicios con ≥1 gate canónico: **{stats['exercises_with_at_least_one_canonical_gate']}** "
        f"({100*stats['exercises_with_at_least_one_canonical_gate']//stats['exercises_total']}%)",
        f"- Ejercicios con 0 gates canónicos: **{stats['exercises_with_zero_canonical_gates']}** "
        f"({100*stats['exercises_with_zero_canonical_gates']//stats['exercises_total']}%)",
        f"- De estos, con ≥1 gate débil (RELATED_TO): **{stats['exercises_with_only_weak_gates']}**",
        "",
        "## Cobertura de gates (¿cuántos gates hay conectados?)",
        "",
        f"- Gates totales en catálogo: **{stats['gates_total']}**",
        f"- Gates referenciados como canónico: **{stats['gates_referenced_by_canonical']}**",
        f"- Gates referenciados solo como débil: **{stats['gates_referenced_only_by_weak']}**",
        f"- Gates huérfanos (nunca referenciados): **{stats['orphan_gates_never_referenced']}**",
        "",
        "## Relaciones consumidas",
        "",
    ]
    for rtype, n in sorted(stats["relations_consumed"].items()):
        marker = "canónica" if rtype in GATE_LINK_RELATIONS else "débil"
        md_lines.append(f"- `{rtype}` ({marker}): {n}")
    md_lines += [
        "",
        "## Distribución de gates canónicos por ejercicio",
        "",
        "| # gates canónicos | # ejercicios |",
        "|---:|---:|",
    ]
    for n, count in stats["distribution_canonical_gates_per_exercise"].items():
        md_lines.append(f"| {n} | {count} |")

    md_lines += [
        "",
        "## Issues detectados",
        "",
        f"- Total issues: **{stats['issues_total']}**",
    ]
    if stats["issues_total"]:
        issue_kinds = Counter(i["kind"] for i in join["issues"])
        for k, n in issue_kinds.most_common():
            md_lines.append(f"  - `{k}`: {n}")

    md_lines += [
        "",
        "## Gates huérfanos (para revisión)",
        "",
        f"Total: **{len(orph)}** gates definidos pero nunca referenciados desde ejercicios.",
        "",
        "```",
        *orph[:50],
    ]
    if len(orph) > 50:
        md_lines.append(f"... y {len(orph) - 50} más (ver `orphan_gates.csv`)")
    md_lines.append("```")
    md_lines += [
        "",
        "## Archivos generados",
        "",
        "- `exercise_gates_mapping.json` — mapping completo (source of truth)",
        "- `exercise_gates_review.csv` — CSV amigable con 582 filas para revisión editorial",
        "- `issues.csv` — issues detectados durante el JOIN",
        "- `orphan_gates.csv` — gates definidos sin referencias",
        "",
        "## Próximo paso",
        "",
        "1. Giuliana revisa `exercise_gates_review.csv` (foco: filas con 0 gates + filas con muchos gates)",
        "2. Giuliana aprueba/ajusta antes de que el mapping se persista en `APP_ExerciseCatalog.gates`",
        "3. Una vez aprobado, la próxima regeneración de v3.1 popula la columna `gates`",
    ]
    (outdir / "REPORT.md").write_text("\n".join(md_lines))

    # 5. Orphan gates CSV
    with (outdir / "orphan_gates.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["gate_id"])
        for gid in orph:
            w.writerow([gid])


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("docs/biblioteca-maestra-v3.1.xlsx"),
        help="Ruta al xlsx v3.1 APP_READY (default: docs/biblioteca-maestra-v3.1.xlsx)",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=Path("docs/rediseno/reports/f13_join_gates"),
        help="Directorio de salida del reporte",
    )
    args = parser.parse_args()

    if not args.source.exists():
        print(f"ERROR: fuente no encontrada: {args.source}", file=sys.stderr)
        return 2

    print(f"Cargando {args.source}...")
    wb = load_workbook_readonly(args.source)

    needed = ["APP_ExerciseCatalog", "APP_GateCatalog", "APP_Relationships"]
    for s in needed:
        if s not in wb.sheetnames:
            print(f"ERROR: falta hoja {s} en {args.source}", file=sys.stderr)
            return 3

    exercises = sheet_to_dicts(wb["APP_ExerciseCatalog"])
    gates = sheet_to_dicts(wb["APP_GateCatalog"])
    relationships = sheet_to_dicts(wb["APP_Relationships"])
    print(f"  {len(exercises)} exercises · {len(gates)} gates · {len(relationships)} relationships")

    print("Construyendo JOIN...")
    join = build_join(exercises, gates, relationships)

    print("Escribiendo reportes...")
    write_reports(join, exercises, args.outdir)

    stats = join["stats"]
    can = stats["exercises_with_at_least_one_canonical_gate"]
    total = stats["exercises_total"]
    weak_only = stats["exercises_with_only_weak_gates"]
    print(
        f"\nCobertura CANÓNICA: {can}/{total} ({100*can//total}%)"
        f" · solo débil: {weak_only}"
    )
    print(f"Gates huérfanos (nunca referenciados): {stats['orphan_gates_never_referenced']}")
    print(f"Issues legítimos (PR/TS→GT): {stats['issues_total']} — documentados como pendientes")
    print(f"\nReporte en: {args.outdir}/REPORT.md")
    return 0


if __name__ == "__main__":
    sys.exit(main())
