#!/usr/bin/env python3
"""
QA_AUDIT extendido · Biblioteca Maestra v3.0

Fase 1 · Task F1.2 · aprobado por Giuliana el 2026-07-15.

Corre sobre `docs/biblioteca-maestra-v3.0-consolidada.xlsx` (fuente única
de curación) y produce reporte JSON+MD. Los checks siguen los 7 vocabularios
canónicos definidos en `docs/rediseno/vocabularios-canonicos.md`.

Política aprobada por Giuliana:
  - Nivel global ESTRICTO: todos los checks técnicos como ERROR.
  - `gates_column_populated`: 100% exigido SOLO en categorías curadas.
    Las no curadas quedan documentadas como pendientes, no bloquean el QA.
  - `dosage_completeness`: ERROR si un ejercicio está Activo con dosis
    totalmente vacía; WARNING si tiene rango parcial.
  - `golden_cases_smoke`: ERROR post-Fase-2 (no aplica en Fase 1).

Uso:
  python scripts/rediseno/qa_audit_v30.py
    [--source docs/biblioteca-maestra-v3.0-consolidada.xlsx]
    [--outdir docs/rediseno/reports/f12_qa_audit]
    [--fail-on-error]  # sale con exit code 1 si hay errores

Fail-closed: cualquier check con findings crece el exit code cuando se pasa
`--fail-on-error`. Sin ese flag, siempre sale 0 (útil para inspección).
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any


# ===========================================================================
# VOCABULARIOS CANÓNICOS (fuente: docs/rediseno/vocabularios-canonicos.md v1.0)
# ===========================================================================

CANON_EQUIPMENT_TOKENS = {
    "gym", "hangboard", "campus", "weights", "rock",
    "home", "bands", "pullup_bar", "trx",
}

CANON_CATEGORY = {
    "fuerza-dedos", "traccion", "antebrazo-muneca-codo", "fuerza-general",
    "resistencia-fuerza", "power-endurance", "hombros-escapulas", "core",
    "tecnica-escalada", "calentamiento", "recuperacion", "campus-potencia",
    "movilidad", "prevencion", "mental",
}

CANON_SEVERITY = {"critical", "high", "medium", "low"}

CANON_ACTION = {
    "STOP_SESSION", "BLOCK", "MANUAL_REVIEW", "STOP_OR_REGRESS",
    "REGRESS", "ADJUST_VOLUME", "HOLD", "SUBSTITUTE", "DEPRIORITIZE",
    "REQUIRE_TECHNIQUE_DRILL", "REORDER", "CUE_ONLY", "ALLOW_ONE_VARIABLE_ONLY",
}

CANON_STRUCTURE = {"warmup", "main", "finisher", "cooldown", "standalone"}

CANON_RISK_LEVEL = {"low", "low-medium", "medium", "medium-high", "high"}

# Prefijos de ID reconocidos por el catálogo — cualquier prefijo distinto
# en `risk_level` etc. es cross-column contamination.
KNOWN_ID_PREFIXES = ("EX-", "PR-", "TS-", "TST-", "GT-", "SRC-", "CAT-")


# ===========================================================================
# HELPERS
# ===========================================================================

def load_workbook_readonly(path: Path):
    """Carga xlsx en modo read_only (bypasea el bug RoadmapStatusTable)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl no está instalado. Correr desde el venv del scratchpad.")
    return load_workbook(path, data_only=True, read_only=True)


def sheet_to_dicts(ws) -> list[dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return []
    headers = [str(h) if h is not None else "" for h in headers]
    out: list[dict[str, str]] = []
    for r in rows_iter:
        if not any(c is not None for c in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            if i < len(r) and r[i] is not None:
                row[h] = str(r[i])
            else:
                row[h] = ""
        out.append(row)
    return out


def sheet_headers(ws) -> list[str]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return []
    return [str(h) if h is not None else "" for h in headers]


# ===========================================================================
# CHECKS
# ===========================================================================

def check_xlsx_integrity(source: Path) -> dict[str, Any]:
    """Detecta el bug de RoadmapStatusTable + tablas nombradas duplicadas.
    Se ejecuta abriendo el archivo en modo NORMAL (no read_only) y captura
    el ValueError característico. Aprobado por Giuliana como check nuevo
    tras encontrar el bug 2026-07-15."""
    from openpyxl import load_workbook
    findings = []
    try:
        load_workbook(source, data_only=True, read_only=False)
    except ValueError as e:
        msg = str(e)
        if "Table with name" in msg:
            findings.append({
                "kind": "duplicate_table_name",
                "detail": msg,
                "location": "workbook-level",
            })
        else:
            findings.append({"kind": "workbook_load_error", "detail": msg})
    return {
        "check_id": "xlsx_integrity_check",
        "level": "error",
        "findings": findings,
        "count": len(findings),
    }


def check_casing_normalization(
    sheet: str, col: str, rows: list[dict[str, str]]
) -> list[dict[str, str]]:
    """Detecta grupos donde los valores difieren SOLO en casing."""
    findings = []
    values = [r.get(col, "").strip() for r in rows if r.get(col, "").strip()]
    if not values:
        return findings
    groups: dict[str, list[str]] = defaultdict(list)
    for v in values:
        groups[v.lower()].append(v)
    for lc, variants in groups.items():
        distinct = sorted(set(variants))
        if len(distinct) > 1:
            findings.append({
                "kind": "casing_duplicate",
                "sheet": sheet,
                "column": col,
                "canonical_key": lc,
                "variants": "|".join(distinct),
                "count_by_variant": ", ".join(
                    f"{v}={variants.count(v)}" for v in distinct
                ),
            })
    return findings


def check_domain(
    sheet: str, col: str, canon: set[str], rows: list[dict[str, str]],
    case_sensitive: bool = True,
) -> list[dict[str, str]]:
    """Valores fuera del vocabulario canónico."""
    findings = []
    off = Counter()
    row_ids = defaultdict(list)
    for r in rows:
        v = r.get(col, "").strip()
        if not v:
            continue
        v_check = v if case_sensitive else v.lower()
        canon_check = canon if case_sensitive else {c.lower() for c in canon}
        if v_check not in canon_check:
            off[v] += 1
            ex_id = r.get("exercise_id") or r.get("gate_id") or r.get("protocol_id") or r.get("id") or "?"
            if len(row_ids[v]) < 3:
                row_ids[v].append(ex_id)
    for v, n in off.most_common(30):
        findings.append({
            "kind": "off_canon",
            "sheet": sheet,
            "column": col,
            "value": v,
            "count": str(n),
            "example_ids": ", ".join(row_ids[v]),
        })
    return findings


def check_cross_column_contamination(
    sheet: str, col: str, rows: list[dict[str, str]]
) -> list[dict[str, str]]:
    """Detecta patrones específicos de contaminación cross-column.
    Reglas conocidas:
      * risk_level con IDs de protocolo (PR-*, TST-*)
      * equipment con prosa (`;` + palabras largas típicas de otras columnas)
      * level_min / level_max con environment (Casa/Gimnasio, Roca, etc)
      * action en Gates con prosa (contiene punto final o >30 chars)
    """
    findings = []
    for r in rows:
        v = r.get(col, "").strip()
        if not v:
            continue
        ex_id = r.get("exercise_id") or r.get("gate_id") or r.get("protocol_id") or r.get("id") or "?"

        # regla 1: risk_level con ID de protocolo/otro tipo
        if col == "risk_level":
            for pfx in ("PR-", "TST-", "TS-", "GT-", "SRC-", "EX-"):
                if v.startswith(pfx):
                    findings.append({
                        "kind": "risk_level_has_id",
                        "sheet": sheet,
                        "column": col,
                        "value": v,
                        "row_id": ex_id,
                    })
                    break

        # regla 2: equipment con prosa larga o oración
        if col == "equipment" and (";" in v or ("." in v and len(v) > 40)):
            findings.append({
                "kind": "equipment_has_prose",
                "sheet": sheet,
                "column": col,
                "value": v[:80],
                "row_id": ex_id,
            })

        # regla 3: level_min / level_max con environment
        if col in {"level_min", "level_max"}:
            env_markers = ("Casa", "Gimnasio", "Roca", "casa", "gimnasio", "roca")
            if any(m in v for m in env_markers):
                findings.append({
                    "kind": "level_has_environment",
                    "sheet": sheet,
                    "column": col,
                    "value": v,
                    "row_id": ex_id,
                })

        # regla 4: action con prosa (punto final, muy larga, o mayúsculas y minúsculas mezcladas)
        if col == "action":
            if v.endswith(".") or len(v) > 40:
                findings.append({
                    "kind": "action_has_prose",
                    "sheet": sheet,
                    "column": col,
                    "value": v[:80],
                    "row_id": ex_id,
                })
    return findings


def check_action_vocabulary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Chequeo especializado del vocabulario de actions en Gates.
    Distingue casing errors de vocabulario off (por semántica de reporte)."""
    findings = []
    canon_lc = {c.lower(): c for c in CANON_ACTION}
    off = Counter()
    for r in rows:
        v = r.get("action", "").strip()
        if not v:
            continue
        if v in CANON_ACTION:
            continue
        if v.lower() in canon_lc:
            # Ejemplo: "block" vs "BLOCK" — es casing, cae en casing_normalization
            continue
        off[v] += 1
    for v, n in off.most_common(50):
        findings.append({
            "kind": "action_off_vocab",
            "sheet": "Gates",
            "column": "action",
            "value": v[:100],
            "count": str(n),
        })
    return findings


def check_subcategory_removed(headers_by_sheet: dict[str, list[str]]) -> list[dict[str, str]]:
    """La columna `subcategory` debe estar eliminada del schema."""
    findings = []
    for sheet, hdrs in headers_by_sheet.items():
        if "subcategory" in hdrs:
            findings.append({
                "kind": "subcategory_present",
                "sheet": sheet,
                "detail": "columna subcategory sigue existiendo · debería estar eliminada",
            })
    return findings


def check_gates_column_populated(
    exercises: list[dict[str, str]], curated_categories: set[str]
) -> list[dict[str, str]]:
    """Filas con `gates` vacío en categorías CURADAS.
    Categorías no-curadas quedan documentadas como pendientes (no error)."""
    findings = []
    pending_by_cat = Counter()
    for r in exercises:
        gates_v = r.get("gates", "").strip()
        cat = r.get("category", "").strip()
        is_empty = (not gates_v) or gates_v.lower().startswith("no especificad")
        if not is_empty:
            continue
        if cat in curated_categories:
            findings.append({
                "kind": "gates_missing_in_curated_category",
                "sheet": "APP_ExerciseCatalog",
                "row_id": r.get("exercise_id", "?"),
                "category": cat,
            })
        else:
            pending_by_cat[cat] += 1
    # Los pendientes se reportan en el mismo bloque como INFO (no findings)
    return findings + [{
        "kind": "pending_non_curated_category",
        "sheet": "APP_ExerciseCatalog",
        "category": cat,
        "count": str(n),
        "level_hint": "info",
    } for cat, n in pending_by_cat.most_common()]


def check_dosage_completeness(protocols: list[dict[str, str]]) -> list[dict[str, str]]:
    """Política de rango aprobada por Giuliana 2026-07-15:
      * ERROR si un protocolo está Activo con dosis TOTALMENTE vacía.
      * WARNING si tiene rango parcial (solo algunos campos).
    """
    findings = []
    dose_cols = ["work_interval", "rest_interval", "sets", "reps", "intensity"]
    for r in protocols:
        status = r.get("app_status", "").strip().lower()
        # Determinar si es "activo" (permite estar en el plan del usuario)
        is_active = ("mvp" in status or "producción" in status or "activo" in status
                     or "lista" in status) and "manual" not in status and "bloque" not in status
        pop = [c for c in dose_cols if r.get(c, "").strip()
               and "no especificad" not in r.get(c, "").lower()]
        pid = r.get("protocol_id", "?")
        if not pop and is_active:
            findings.append({
                "kind": "active_no_dosage",
                "sheet": "Protocols",
                "row_id": pid,
                "status": r.get("app_status", ""),
                "detail": "Activo sin ningún campo de dosis poblado",
                "level": "error",
            })
        elif pop and len(pop) < len(dose_cols):
            findings.append({
                "kind": "partial_dosage",
                "sheet": "Protocols",
                "row_id": pid,
                "populated": ", ".join(pop),
                "missing": ", ".join(c for c in dose_cols if c not in pop),
                "level": "warning",
            })
    return findings


def check_relationships_orphans(
    relationships: list[dict[str, str]],
    known_ids: dict[str, set[str]],
) -> list[dict[str, str]]:
    """from_id y to_id deben corresponder a IDs existentes según su prefijo."""
    findings = []
    known_all = set()
    for _pfx, ids in known_ids.items():
        known_all |= ids
    orphan_from = 0
    orphan_to = 0
    sample_orph_from = []
    sample_orph_to = []
    for r in relationships:
        fid = r.get("from_id", "").strip()
        tid = r.get("to_id", "").strip()
        rid = r.get("relationship_id", "?")
        if fid and fid not in known_all:
            orphan_from += 1
            if len(sample_orph_from) < 20:
                sample_orph_from.append(f"{rid}: {fid}")
        if tid and tid not in known_all:
            orphan_to += 1
            if len(sample_orph_to) < 20:
                sample_orph_to.append(f"{rid}: {tid}")
    if orphan_from:
        findings.append({
            "kind": "from_id_orphan",
            "sheet": "Relationships",
            "count": str(orphan_from),
            "sample": " | ".join(sample_orph_from),
        })
    if orphan_to:
        findings.append({
            "kind": "to_id_orphan",
            "sheet": "Relationships",
            "count": str(orphan_to),
            "sample": " | ".join(sample_orph_to),
        })
    return findings


# ===========================================================================
# RUNNER
# ===========================================================================

def determine_curated_categories(exercises: list[dict[str, str]]) -> set[str]:
    """Una categoría se considera curada si ≥50% de sus ejercicios tienen
    `gates` no vacío. Umbral pragmático — Giuliana aprobó 100% para las
    categorías curadas pero la definición de "curada" es empírica hasta que
    haya un flag explícito en el catálogo."""
    cat_stats: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "with_gates": 0})
    for r in exercises:
        cat = r.get("category", "").strip()
        if not cat:
            continue
        cat_stats[cat]["total"] += 1
        gates_v = r.get("gates", "").strip()
        if gates_v and not gates_v.lower().startswith("no especificad"):
            cat_stats[cat]["with_gates"] += 1
    curated = {
        cat for cat, s in cat_stats.items()
        if s["total"] > 0 and s["with_gates"] / s["total"] >= 0.5
    }
    return curated


def run_audit(source: Path) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []

    # Check 1: xlsx integrity (abre el file en modo normal)
    checks.append(check_xlsx_integrity(source))

    wb = load_workbook_readonly(source)

    # Cargar hojas relevantes (usar nombres del v3.0 CONSOLIDADA)
    def try_load(name: str) -> list[dict[str, str]]:
        if name in wb.sheetnames:
            return sheet_to_dicts(wb[name])
        return []

    exercises_v30 = try_load("Exercises")
    protocols_v30 = try_load("Protocols")
    tests_v30 = try_load("Tests")
    gates_v30 = try_load("Gates")
    sources_v30 = try_load("Sources")
    relationships_v30 = try_load("Relationships")

    # También intentar v3.1-style names (por si Giuliana corre esto sobre v3.1)
    exercises_v31 = try_load("APP_ExerciseCatalog")
    protocols_v31 = try_load("APP_ProtocolCatalog")
    gates_v31 = try_load("APP_GateCatalog")
    relationships_v31 = try_load("APP_Relationships")

    exercises = exercises_v31 or exercises_v30
    protocols = protocols_v31 or protocols_v30
    gates = gates_v31 or gates_v30
    relationships = relationships_v31 or relationships_v30

    headers_by_sheet = {
        "Exercises": list(exercises[0].keys()) if exercises else [],
        "Protocols": list(protocols[0].keys()) if protocols else [],
        "Gates": list(gates[0].keys()) if gates else [],
    }

    # Check 2: subcategory removed
    checks.append({
        "check_id": "subcategory_removed",
        "level": "error",
        "findings": check_subcategory_removed(headers_by_sheet),
        "count": len(check_subcategory_removed(headers_by_sheet)),
    })

    # Check 3: casing normalization en columnas clave
    casing_cases = [
        ("Exercises", "risk_level", exercises),
        ("Exercises", "category", exercises),
        ("Exercises", "equipment", exercises),
        ("Gates", "severity", gates),
        ("Gates", "action", gates),
        ("Protocols", "risk_level", protocols),
    ]
    casing_findings = []
    for sheet, col, rows in casing_cases:
        casing_findings.extend(check_casing_normalization(sheet, col, rows))
    checks.append({
        "check_id": "casing_normalization",
        "level": "error",
        "findings": casing_findings,
        "count": len(casing_findings),
    })

    # Check 4-9: domain checks
    domain_cases = [
        ("domain_check_category", "Exercises", "category", CANON_CATEGORY, exercises, True),
        ("domain_check_severity", "Gates", "severity", CANON_SEVERITY, gates, False),
        ("domain_check_structure", "Protocols", "structure", CANON_STRUCTURE, protocols, False),
        ("domain_check_risk_level_exercises", "Exercises", "risk_level", CANON_RISK_LEVEL, exercises, False),
        ("domain_check_risk_level_protocols", "Protocols", "risk_level", CANON_RISK_LEVEL, protocols, False),
    ]
    for check_id, sheet, col, canon, rows, cs in domain_cases:
        f = check_domain(sheet, col, canon, rows, case_sensitive=cs)
        checks.append({"check_id": check_id, "level": "error", "findings": f, "count": len(f)})

    # Check equipment: la columna en Exercises es texto libre, no un token.
    # Chequeamos que si viene un array (separado por `;` o `,`), cada token
    # esté en CANON_EQUIPMENT_TOKENS. Si es prosa, cae en cross_column.
    equip_off = []
    for r in exercises:
        v = r.get("equipment", "").strip()
        if not v or v.lower().startswith("no especificad"):
            continue
        # Ignorar si ya cayó en cross_column (prosa larga con `;` o punto)
        if ";" in v or ("." in v and len(v) > 40):
            continue
        # Tokenizar por `,` `/` o espacios cuando es simple
        tokens = [t.strip().lower() for t in re.split(r"[,/]", v) if t.strip()]
        for tok in tokens:
            if tok not in CANON_EQUIPMENT_TOKENS:
                equip_off.append({
                    "kind": "off_canon",
                    "sheet": "Exercises",
                    "column": "equipment",
                    "value": v[:80],
                    "off_token": tok,
                    "row_id": r.get("exercise_id", "?"),
                })
                break
    checks.append({
        "check_id": "domain_check_equipment",
        "level": "error",
        "findings": equip_off[:200],
        "count": len(equip_off),
    })

    # Check 10: action_vocabulary_check
    action_findings = check_action_vocabulary(gates)
    checks.append({
        "check_id": "action_vocabulary_check",
        "level": "error",
        "findings": action_findings,
        "count": len(action_findings),
    })

    # Check 11: cross_column_contamination
    cross_col_cases = [
        ("Exercises", "risk_level", exercises),
        ("Exercises", "equipment", exercises),
        ("Exercises", "level_min", exercises),
        ("Exercises", "level_max", exercises),
        ("Gates", "action", gates),
    ]
    cross_findings = []
    for sheet, col, rows in cross_col_cases:
        cross_findings.extend(check_cross_column_contamination(sheet, col, rows))
    checks.append({
        "check_id": "cross_column_contamination",
        "level": "error",
        "findings": cross_findings,
        "count": len(cross_findings),
    })

    # Check 12: gates_column_populated (solo en categorías curadas)
    curated = determine_curated_categories(exercises)
    gates_pop = check_gates_column_populated(exercises, curated)
    real_findings = [f for f in gates_pop if f.get("kind") != "pending_non_curated_category"]
    pending_findings = [f for f in gates_pop if f.get("kind") == "pending_non_curated_category"]
    checks.append({
        "check_id": "gates_column_populated",
        "level": "error",
        "curated_categories": sorted(curated),
        "findings": real_findings,
        "count": len(real_findings),
        "pending_non_curated": pending_findings,
    })

    # Check 13: dosage_completeness
    dose_findings = check_dosage_completeness(protocols)
    dose_errors = [f for f in dose_findings if f.get("level") == "error"]
    dose_warnings = [f for f in dose_findings if f.get("level") == "warning"]
    checks.append({
        "check_id": "dosage_completeness",
        "level": "error",
        "findings": dose_errors,
        "count": len(dose_errors),
        "warnings": dose_warnings,
        "warning_count": len(dose_warnings),
    })

    # Check 14: relationships orphans
    known_ids = {
        "EX-": {r.get("exercise_id", "") for r in exercises if r.get("exercise_id", "").startswith("EX-")},
        "PR-": {r.get("protocol_id", "") for r in protocols if r.get("protocol_id", "").startswith("PR-")},
        "TS-": {r.get("test_id", "") for r in tests_v30 if r.get("test_id", "").startswith("TS-")},
        "TST-": {r.get("test_id", "") for r in tests_v30 if r.get("test_id", "").startswith("TST-")},
        "GT-": {r.get("gate_id", "") for r in gates if r.get("gate_id", "").startswith("GT-")},
        "SRC-": {r.get("source_id", "") for r in sources_v30 if r.get("source_id", "").startswith("SRC-")},
        "CAT-": set(),  # categorías como entidad — v3.0 no las tiene como IDs propios
    }
    orph_findings = check_relationships_orphans(relationships, known_ids)
    checks.append({
        "check_id": "relationships_orphans",
        "level": "error",
        "findings": orph_findings,
        "count": len(orph_findings),
    })

    # Golden cases smoke: pendiente para post-Fase-2
    checks.append({
        "check_id": "golden_cases_smoke",
        "level": "error",
        "findings": [],
        "count": 0,
        "note": "PENDIENTE post-Fase-2. GC-001 a GC-007 deben producir su expected_focus.",
    })

    # Resumen agregado
    total_errors = sum(c["count"] for c in checks if c.get("level") == "error")
    total_warnings = sum(c.get("warning_count", 0) for c in checks)

    return {
        "source": str(source),
        "checks": checks,
        "totals": {
            "checks_run": len(checks),
            "errors": total_errors,
            "warnings": total_warnings,
            "curated_categories": sorted(curated),
        },
    }


def write_report(audit: dict[str, Any], outdir: Path) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "audit.json").write_text(
        json.dumps(audit, indent=2, ensure_ascii=False, sort_keys=True)
    )

    md_lines = [
        "# QA_AUDIT extendido · Biblioteca Maestra v3.0",
        "",
        f"**Fuente**: `{audit['source']}`",
        f"**Fecha**: 2026-07-15 · Fase 1 · Task F1.2",
        f"**Política aprobada**: nivel estricto · todos los checks técnicos como ERROR",
        "",
        f"## Resumen",
        "",
        f"- Checks corridos: **{audit['totals']['checks_run']}**",
        f"- Errores totales: **{audit['totals']['errors']}**",
        f"- Warnings totales: **{audit['totals']['warnings']}**",
        f"- Categorías curadas (≥50% con gates): **{len(audit['totals']['curated_categories'])}**",
        "",
        "## Categorías curadas detectadas",
        "",
    ]
    for c in audit["totals"]["curated_categories"]:
        md_lines.append(f"- `{c}`")

    md_lines.append("")
    md_lines.append("## Detalle por check")
    md_lines.append("")

    for c in audit["checks"]:
        status = "✅ OK" if c["count"] == 0 else f"❌ {c['count']} findings"
        md_lines.append(f"### {c['check_id']} · {c['level']} · {status}")
        md_lines.append("")
        if c.get("note"):
            md_lines.append(f"> {c['note']}")
            md_lines.append("")
        # Sample de findings
        if c["findings"]:
            sample = c["findings"][:15]
            md_lines.append("Sample:")
            md_lines.append("")
            for f in sample:
                # Formatear finding como bullet compacto
                bits = []
                for k, v in f.items():
                    if k == "level_hint":
                        continue
                    bits.append(f"`{k}={v}`")
                md_lines.append(f"- {' · '.join(bits)}")
            if len(c["findings"]) > 15:
                md_lines.append(f"- ... y {len(c['findings']) - 15} más")
            md_lines.append("")
        # Warnings anexos
        if c.get("warnings"):
            md_lines.append(f"**Warnings ({len(c['warnings'])}):**")
            md_lines.append("")
            for w in c["warnings"][:10]:
                bits = [f"`{k}={v}`" for k, v in w.items() if k != "level"]
                md_lines.append(f"- {' · '.join(bits)}")
            if len(c["warnings"]) > 10:
                md_lines.append(f"- ... y {len(c['warnings']) - 10} más")
            md_lines.append("")
        # Pendientes (info) para gates_column_populated
        if c.get("pending_non_curated"):
            md_lines.append(f"**Pendientes (categorías no-curadas, info · no bloquean):**")
            md_lines.append("")
            for p in c["pending_non_curated"][:30]:
                md_lines.append(f"- `{p['category']}` · {p['count']} ejercicios sin gate")
            md_lines.append("")

    (outdir / "REPORT.md").write_text("\n".join(md_lines))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path,
                        default=Path("docs/biblioteca-maestra-v3.0-consolidada.xlsx"))
    parser.add_argument("--outdir", type=Path,
                        default=Path("docs/rediseno/reports/f12_qa_audit"))
    parser.add_argument("--fail-on-error", action="store_true")
    args = parser.parse_args()

    if not args.source.exists():
        print(f"ERROR: fuente no encontrada: {args.source}", file=sys.stderr)
        return 2

    print(f"Cargando {args.source}...")
    audit = run_audit(args.source)
    print("Escribiendo reporte...")
    write_report(audit, args.outdir)

    t = audit["totals"]
    print(f"\nQA_AUDIT: {t['errors']} errors · {t['warnings']} warnings · "
          f"{len(t['curated_categories'])} categorías curadas")
    print(f"Reporte: {args.outdir}/REPORT.md")

    if args.fail_on_error and t["errors"] > 0:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
