#!/usr/bin/env python3
"""
Regenerador v3.0 (maestro) → v3.1 (APP_READY) · Fase 1 · Task F1.4.

Aprobado por Giuliana el 2026-07-15/16 con las 4 confirmaciones:
  1. Dosis con parser — Giuliana cura formato natural ("3–15 s",
     "70–80%; buffer"), este script extrae rangos numéricos + preserva
     matices en notes_dosage ("buffer", "por rep", "según fases").
  2. gates_column_populated — el check corre sobre v3.1 regenerada.
  3. dosage_completeness — heurística para arrancar + promoción a
     `manual_review` cuando la dosis está totalmente vacía.
  4. Arrancar en frío, revisar sobre v3.1 los residuales + el CSV de
     contaminación (44 filas).

## Qué hace

1. Carga v3.0 (hojas Exercises, Protocols, Tests, Gates, Sources,
   Relationships + spec sheets ProfileSchema, SelectionPipeline).
2. Aplica los 8 vocabularios canónicos vía mapping tables ampliables.
3. Parsea dosis (Protocols → rangos numéricos + notes_dosage).
4. Detecta contaminación cross-column y separa a `data_quality_issues.csv`.
5. Expande Relationships (`; ` → filas individuales).
6. Aplana Exercises (46 → 21 cols), Protocols (34 → 22).
7. Genera `APP_OnboardingMap` desde ProfileSchema.
8. Genera `APP_SessionBuilderInputs` desde SelectionPipeline.
9. Escribe xlsx v3.1 nuevo y ejecuta `qa_audit_v30.py` sobre él.

## Fail-closed

- Cualquier fila que no se pueda mapear a un vocabulario canónico se emite
  igual (con valor crudo) pero se registra en `unmapped_values.csv` para
  ampliar la mapping table en el próximo sprint.
- Filas con contaminación cross-column detectada se emiten al v3.1 pero
  con flag `data_quality_flag=1` y detalle en el CSV separado.
- El script NO edita el v3.0 — solo lee. La curación queda en Giuliana.

Uso:
    python scripts/rediseno/regenerate_v31.py
      [--source docs/biblioteca-maestra-v3.0-consolidada.xlsx]
      [--target docs/biblioteca-maestra-v3.1.xlsx]
      [--outdir docs/rediseno/reports/f14_regenerate]
      [--skip-qa]  # no correr el QA post-regeneración
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("openpyxl no está instalado. Correr desde el venv del scratchpad.")


# ===========================================================================
# VOCABULARIOS CANÓNICOS + MAPPING TABLES
# Fuente: docs/rediseno/vocabularios-canonicos.md v1.0
# ===========================================================================

CANON_EQUIPMENT_TOKENS = {
    "gym", "hangboard", "campus", "weights", "rock",
    "home", "bands", "pullup_bar", "trx",
    # +2 tokens aprobados por Giuliana 2026-07-16 (F1.5)
    "dynamometer", "pinch_block",
    # +2 tokens aprobados por Giuliana 2026-07-20 (curacion-bulk-11cat-v1)
    "campus_board",  # escalera bachar · campus específico como equipo
    "ball",          # balón medicinal / pelota blanda / fitball
}

CANON_CATEGORY = {
    "fuerza-dedos", "traccion", "antebrazo-muneca-codo", "fuerza-general",
    "resistencia-fuerza", "power-endurance", "hombros-escapulas", "core",
    "tecnica-escalada", "calentamiento", "recuperacion", "campus-potencia",
    "movilidad", "prevencion", "mental",
}

CANON_SEVERITY = {"critical", "high", "medium", "low"}
CANON_STRUCTURE = {"warmup", "main", "finisher", "cooldown", "standalone"}
CANON_RISK_LEVEL = {"low", "low-medium", "medium", "medium-high", "high"}
CANON_STATUS = {"active", "manual_review", "draft", "deprecated"}
CANON_ACTION = {
    "STOP_SESSION", "BLOCK", "MANUAL_REVIEW", "STOP_OR_REGRESS",
    "REGRESS", "ADJUST_VOLUME", "HOLD", "SUBSTITUTE", "DEPRIORITIZE",
    "REQUIRE_TECHNIQUE_DRILL", "REORDER", "CUE_ONLY", "ALLOW_ONE_VARIABLE_ONLY",
}

# ---------------------------------------------------------------------------
# Mapping tables · valor crudo (lowercased) → canónico
# Ampliar aquí en sprints futuros a medida que aparezcan variantes nuevas.
# ---------------------------------------------------------------------------

CATEGORY_ALIASES: dict[str, str] = {
    "fuerza de dedos / hangboard": "fuerza-dedos",
    "fuerza de dedos/hangboard": "fuerza-dedos",
    "tracción y dominadas": "traccion",
    "traccion y dominadas": "traccion",
    "antebrazo, muñeca y codo": "antebrazo-muneca-codo",
    "antebrazo/muñeca": "antebrazo-muneca-codo",
    "antebrazo": "antebrazo-muneca-codo",
    "muñeca": "antebrazo-muneca-codo",
    "muñeca / dedos": "antebrazo-muneca-codo",
    "fuerza general / piernas / acondicionamiento": "fuerza-general",
    "fuerza general": "fuerza-general",
    "resistencia de fuerza": "resistencia-fuerza",
    "power endurance": "power-endurance",
    "hombros / escápulas": "hombros-escapulas",
    "escápulas / hombros": "hombros-escapulas",
    "hombros / manguito rotador": "hombros-escapulas",
    "hombros / movilidad activa": "hombros-escapulas",
    "hombro": "hombros-escapulas",
    "hombro/escápula": "hombros-escapulas",
    "hombro / escápula": "hombros-escapulas",
    "escápula": "hombros-escapulas",
    "hombro posterior": "hombros-escapulas",
    "hombro/torácica": "hombros-escapulas",
    "hombro/pectoral": "hombros-escapulas",
    "pecho/hombro": "hombros-escapulas",
    "dorsal/hombro": "hombros-escapulas",
    "cervical/trapecio": "hombros-escapulas",
    "cervical": "hombros-escapulas",
    "torácica/hombro posterior": "hombros-escapulas",
    "core": "core",
    "técnica de escalada": "tecnica-escalada",
    "tecnica de escalada": "tecnica-escalada",
    "arc": "tecnica-escalada",
    "arc/técnica": "tecnica-escalada",
    "arc/tecnica": "tecnica-escalada",
    "arc/táctica": "tecnica-escalada",
    "calentamientos": "calentamiento",
    "calentamiento": "calentamiento",
    "recuperación": "recuperacion",
    "recuperacion": "recuperacion",
    "campus y potencia específica": "campus-potencia",
    "campus y potencia especifica": "campus-potencia",
    "movilidad": "movilidad",
    "movilidad activa": "movilidad",
    "movilidad anterior": "movilidad",
    "movilidad torácica": "movilidad",
    "cadera": "movilidad",
    "cadera/aductores": "movilidad",
    "cadera/isquios": "movilidad",
    "cadera/torácica": "movilidad",
    "cadera/tobillo": "movilidad",
    "cadera/tobillo/isquios": "movilidad",
    "cadera/dorsal": "movilidad",
    "isquios": "movilidad",
    "isquios/espalda": "movilidad",
    "aductores": "movilidad",
    "aductores/isquios": "movilidad",
    "columna": "movilidad",
    "columna torácica": "movilidad",
    "torácica/cadera/hombro": "movilidad",
    "cadena posterior": "movilidad",
    "cadena posterior/hombros": "movilidad",
    "cuádriceps/flexor cadera": "movilidad",
    "glúteo/rotadores cadera": "movilidad",
    "tobillo": "movilidad",
    "tobillo/gemelo": "movilidad",
    "tobillo/sóleo": "movilidad",
    "pie/tobillo": "movilidad",
    "tríceps/dorsal": "movilidad",
    "dorsal/torácica": "movilidad",
    "prevención de lesiones general": "prevencion",
    "prevencion de lesiones general": "prevencion",
    "trabajo mental": "mental",
    "global": "mental",
    "periodización y programación": "mental",  # meta-conceptual, va aquí interim
    "tests y evaluaciones integradas": "mental",
}

# Equipment · substrings → tokens canónicos. Se aplica en orden (long
# match first para evitar que "banda" capture "banda/palo").
EQUIPMENT_SUBSTRING_MAP: list[tuple[str, list[str]]] = [
    # Ajustes F1.5 aprobados por Giuliana 2026-07-16
    ("dinamómetro específico", ["dynamometer"]),
    ("dinamómetro", ["dynamometer"]),
    ("dynamometer", ["dynamometer"]),
    ("pinch ball/block", ["pinch_block"]),
    ("pinch block", ["pinch_block"]),
    ("bloque de pinza", ["pinch_block"]),
    ("borde estable", ["hangboard"]),
    ("system wall", ["gym"]),
    # Curacion-bulk-11cat-v1 (Giuliana 2026-07-20) · REGLA C1.
    # Mapear herramientas reales fuera de catálogo a canon.
    ("escalera bachar", ["campus_board"]),
    ("campus board", ["campus_board"]),
    ("campus", ["campus_board"]),  # colapsa "campus" crudo → equipo específico
    ("balón medicinal", ["ball"]),
    ("balon medicinal", ["ball"]),
    ("pelota blanda", ["ball"]),
    ("fitball", ["ball"]),
    ("pelota", ["ball"]),          # genérico al final para no capturar antes
    ("cubeta con arroz", ["home"]),
    ("mesa", ["home"]),
    ("trineo", ["gym"]),
    ("bicicleta", ["gym"]),
    ("caminadora", ["gym"]),
    ("terreno", ["gym"]),
    ("exterior", ["gym"]),
    ("cuerda", ["gym"]),
    # curacion-dedos-v1 (Giuliana 2026-07-19): EX-FIN-033 "Bloque/volumen"
    # es un bloque de pinza (misma familia que EX-FIN-031). No es system
    # board — el regen v1 lo apuntaba a gym por error.
    ("bloque/volumen", ["pinch_block"]),
    ("bloque / volumen", ["pinch_block"]),
    ("board", ["gym"]),  # spray/moon/kilter boards
    ("liga elástica", ["bands"]),
    ("liga elastica", ["bands"]),
    ("sin equipo", ["home"]),
    ("no equipment", ["home"]),
    ("bodyweight only", ["home"]),
    # Combos primero (longer match wins)
    ("hangboard/pocket board", ["hangboard"]),
    ("hangboard con lastre", ["hangboard", "weights"]),
    ("hangboard + lastre", ["hangboard", "weights"]),
    ("hangboard + banda", ["hangboard", "bands"]),
    ("hangboard + polea", ["hangboard", "bands"]),
    ("hangboard con bordes variables", ["hangboard"]),
    ("hangboard o borde estable", ["hangboard"]),
    ("hangboard/pocket", ["hangboard"]),
    ("pocket board", ["hangboard"]),
    ("hangboard", ["hangboard"]),
    ("fingerboard", ["hangboard"]),
    ("campus board", ["campus", "gym"]),
    ("campus", ["campus", "gym"]),
    ("muro de boulder", ["gym"]),
    ("muro de escalada", ["gym"]),
    ("muro/roca", ["gym", "rock"]),
    ("muro/colchoneta", ["gym", "home"]),
    ("muro / colchoneta", ["gym", "home"]),
    ("muro", ["gym"]),
    ("boulder wall", ["gym"]),
    ("boulder / ruta", ["gym"]),
    ("gym", ["gym"]),
    ("gimnasio", ["gym"]),
    ("rutas", ["gym"]),
    ("ruta", ["gym"]),
    ("presas", ["gym"]),
    ("travesía", ["gym"]),
    ("roca", ["rock"]),
    ("barra/anillas", ["pullup_bar", "trx"]),
    ("barra/mancuernas", ["pullup_bar", "weights"]),
    ("barra/discos", ["pullup_bar", "weights"]),
    ("barra de dominadas", ["pullup_bar"]),
    ("barra", ["pullup_bar"]),
    ("pullup bar", ["pullup_bar"]),
    ("pull-up bar", ["pullup_bar"]),
    ("dominadas", ["pullup_bar"]),
    ("anillas", ["trx"]),
    ("trx", ["trx"]),
    ("banda/palo", ["bands", "home"]),
    ("banda/toalla", ["bands", "home"]),
    ("banda elástica", ["bands"]),
    ("bandas elásticas", ["bands"]),
    ("resistencia elástica", ["bands"]),
    ("bandas", ["bands"]),
    ("banda", ["bands"]),
    ("polea", ["bands"]),
    ("mancuerna ligera", ["weights"]),
    ("mancuernas/kettlebells", ["weights"]),
    ("mancuernas/barra", ["weights", "pullup_bar"]),
    ("mancuernas", ["weights"]),
    ("mancuerna", ["weights"]),
    ("kettlebells", ["weights"]),
    ("kettlebell", ["weights"]),
    ("discos", ["weights"]),
    ("disco", ["weights"]),
    ("lastre", ["weights"]),
    ("peso corporal", ["home"]),
    ("bodyweight", ["home"]),
    ("peso ligero", ["weights"]),
    ("peso", ["weights"]),
    ("colchoneta", ["home"]),
    ("suelo", ["home"]),
    ("floor", ["home"]),
    ("toalla", ["home"]),
    ("libreta", ["home"]),
    ("papel", ["home"]),
    ("pared/apoyo", ["home"]),
    ("puerta/pared", ["home"]),
    ("pared/banco", ["home"]),
    ("pared", ["home"]),
    ("sofá", ["home"]),
    ("silla", ["home"]),
    ("cajón", ["home"]),
    ("escalón", ["home"]),
    ("foam roller", ["home"]),
    ("banco", ["home"]),
    ("app", ["home"]),
    ("ninguno", ["home"]),
    ("n/a", ["home"]),
    ("no aplica", ["home"]),
    ("apoyo opcional", ["home"]),
]

SEVERITY_ALIASES: dict[str, str] = {
    "crítico": "critical", "crítica": "critical", "critico": "critical",
    "critica": "critical", "critical": "critical", "blocker": "critical",
    "alto": "high", "alta": "high", "high": "high",
    "medio": "medium", "media": "medium", "medium": "medium",
    "bajo": "low", "baja": "low", "low": "low",
    "bajo-medio": "low", "bajo/medio": "low",
    "media-alta": "high",  # decisión conservadora
    "medio-alto": "medium-high",  # solo para risk_level, no severity
    "medio/alto": "medium-high",
    "warning": "medium",  # se marca aparte en is_warning
}

# is_warning: valores crudos que indican flag warning (no severity real).
SEVERITY_IS_WARNING_MARKERS = {"warning", "aviso", "advertencia"}

# risk_level son solo 5 valores canónicos del enum #7. "Regla" y otros
# no-standard van a "medium-high" como fallback conservador (el ejercicio
# igual queda flagged en cross_column_contamination cuando aplica).
RISK_LEVEL_ALIASES: dict[str, str] = {
    "bajo": "low", "low": "low",
    "bajo-medio": "low-medium", "bajo/medio": "low-medium",
    "medio": "medium", "medium": "medium",
    "medio-alto": "medium-high", "medio/alto": "medium-high",
    "alto": "high", "high": "high",
    "regla": "medium-high",  # fallback conservador — errata editorial
}

# Mapa de action normalizada. La familia canonical se preserva; las
# variantes se normalizan a la canónica. Las que no matchean se
# reportan como unmapped y quedan crudas.
ACTION_ALIASES: dict[str, str] = {
    # BLOCK family
    "block": "BLOCK",
    "bloquear": "BLOCK",
    "bloquear.": "BLOCK",
    "block_high_load": "BLOCK",
    "block_loaded_wrist": "BLOCK",
    "block_grip_load": "BLOCK",
    "block_flexor_loading": "BLOCK",
    "block_extensor_loading": "BLOCK",
    "block_pinch": "BLOCK",
    "block_rice_bucket": "BLOCK",
    "block_loaded_carry": "BLOCK",
    "block_next_high_load": "BLOCK",
    "block_until_warmup": "BLOCK",
    "avoid_extra_grip_load": "BLOCK",
    "avoid_fatiguing_forearm": "BLOCK",
    # STOP_SESSION
    "stop": "STOP_SESSION",
    "stop_and_refer": "STOP_SESSION",
    # MANUAL_REVIEW / MANUAL_ONLY
    "manual_only": "MANUAL_REVIEW",
    "manual_review": "MANUAL_REVIEW",
    "require_coach_validation": "MANUAL_REVIEW",
    "require_professional_validation": "MANUAL_REVIEW",
    "professional_validation": "MANUAL_REVIEW",
    "require_pro_validation": "MANUAL_REVIEW",
    "block_and_refer": "MANUAL_REVIEW",
    # STOP_OR_REGRESS
    "stop_or_regress": "STOP_OR_REGRESS",
    "regress_or_block": "STOP_OR_REGRESS",
    "block_or_regress": "STOP_OR_REGRESS",
    # REGRESS
    "regress": "REGRESS",
    "regress_next": "REGRESS",
    "reduce_training_load": "REGRESS",
    "reduce_elbow_flexor_load": "REGRESS",
    "modify_to_submax": "REGRESS",
    # ADJUST_VOLUME
    "adjust_load": "ADJUST_VOLUME",
    "adjust_volume": "ADJUST_VOLUME",
    "modify": "ADJUST_VOLUME",
    # SUBSTITUTE
    "substitute": "SUBSTITUTE",
    "select_bodyweight_or_band": "SUBSTITUTE",
    "change_test": "SUBSTITUTE",
    # HOLD (idem)
    "hold": "HOLD",
    # DEPRIORITIZE
    "deprioritize": "DEPRIORITIZE",
    # REQUIRE_TECHNIQUE_DRILL
    "require_technique_drill": "REQUIRE_TECHNIQUE_DRILL",
    # REORDER
    "reorder": "REORDER",
    "warn_or_reorder": "REORDER",
    # CUE_ONLY
    "cue": "CUE_ONLY",
    "cue_or_regress": "CUE_ONLY",
    "cue_only": "CUE_ONLY",
    # ALLOW_ONE_VARIABLE_ONLY
    "allow_one_variable_only": "ALLOW_ONE_VARIABLE_ONLY",
    # BLOCK_METHODS (scope-specific, se mapea a BLOCK con applies_to)
    "block_methods": "BLOCK",
}

# Structure aliases + heurística substring. Fallback a "main" cuando
# el texto solo dice "no especificado" o similar — mejor default seguro.
STRUCTURE_ALIASES: dict[str, str] = {
    "principal": "main",
    "bloque principal": "main",
    "bloque principal/técnico": "main",
    "bloque principal/técnico; posición exacta no especificada": "main",
    "bloque específico": "main",
    "bloque de fuerza": "main",
    "bloque de fuerza técnica": "main",
    "bloque de resistencia": "main",
    "bloque mixto": "main",
    "primer bloque tras calentamiento": "main",
    "después de calentamiento": "main",
    "día de resistencia": "main",
    "día de resistencia tras calentamiento": "main",
    "día auxiliar": "main",
    "fuerza general": "main",
    "secundario": "finisher",
    "auxiliar": "finisher",
    "calentamiento": "warmup",
    "calentamiento o accesorio": "warmup",
    "warmup": "warmup",
    "técnica/preparación": "warmup",
    "fin de sesión": "cooldown",
    "cooldown": "cooldown",
    "enfriamiento": "cooldown",
    "standalone": "standalone",
    "no especificado en la fuente": "main",
    "requiere validación profesional": "main",
    "no diagnóstico": "main",
    "observacional": "standalone",
    "no test máximo con dolor": "main",
    "no automático con dolor": "main",
    "mvp usa post-check-in": "main",
    "evita hangboard/campus avanzados": "main",
    "requiere gates": "main",
    "roca cuenta como carga": "main",
    "no sustituye evaluación clínica": "main",
    "id referenciado por relaciones pero sin fila definida en catálogos fuente. no us": "draft",
    "después de alta profesional": "main",
    "solo manual": "standalone",
    "solo con profesional": "standalone",
    "complemento / prevención": "finisher",
    "complemento": "finisher",
    "accesorio": "finisher",
    "potencia; no cerca de fatiga": "main",
    "sesión de muro": "standalone",
    "sesión de muro / complemento": "main",
    "antes de escalar/entrenar": "warmup",
    "antes de dinámicos": "warmup",
    "antes de prescripción": "standalone",
    "después de sesión": "cooldown",
    "semana de escalada alta": "main",
    "planificación": "standalone",
}

# Substring heurística para structure — cuando exact match falla,
# probar substrings (aprobado por Giuliana como corto plazo).
STRUCTURE_SUBSTRING_ORDER = [
    ("calentamiento", "warmup"),
    ("warmup", "warmup"),
    ("cooldown", "cooldown"),
    ("enfriamiento", "cooldown"),
    ("fin de sesión", "cooldown"),
    ("finisher", "finisher"),
    ("auxiliar", "finisher"),
    ("secundario", "finisher"),
    ("standalone", "standalone"),
    ("bloque", "main"),
    ("principal", "main"),
    ("día", "main"),
    ("resistencia", "main"),
    ("fuerza", "main"),
    ("técnica", "main"),
    ("no especificad", "main"),
]

# Status heurística: substring → canónico. Aprobado por Giuliana como
# corto plazo — el enum real vive en el siguiente sprint.
STATUS_HEURISTIC_ORDER = [
    ("manual", "manual_review"),
    ("no automático", "manual_review"),
    ("bloqueado", "manual_review"),
    ("investigación", "manual_review"),
    # F4.0 (Sprint 13-24 usan requires_validation): 'Sí' + 'Requiere validación'
    # → manual_review. 'No' → active.
    ("requiere validación", "manual_review"),
    ("requiere validacion", "manual_review"),
    ("sí", "manual_review"),
    ("borrador", "draft"),
    ("draft", "draft"),
    ("deprecated", "deprecated"),
    ("retirado", "deprecated"),
    ("mvp", "active"),
    ("producción", "active"),
    ("produccion", "active"),
    ("activo", "active"),
    ("active", "active"),
    ("lista", "active"),
    ("validado", "active"),
    ("listo", "active"),
    # 'No' solo (de requires_validation=No en sprints) → active
    ("no", "active"),
]


# ===========================================================================
# HELPERS DE CANONICALIZACIÓN
# ===========================================================================

def norm_lower(v: str) -> str:
    return (v or "").strip().lower()


def canon_category(v: str) -> tuple[str, bool]:
    """Devuelve (canonical, mapped_ok). Si no matchea, deja crudo."""
    if not v.strip():
        return "", True
    k = norm_lower(v)
    if k in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[k], True
    if k in CANON_CATEGORY:
        return k, True
    return v.strip(), False


def canon_severity(v: str) -> tuple[str, bool, bool]:
    """Devuelve (canonical, mapped_ok, is_warning)."""
    if not v.strip():
        return "", True, False
    k = norm_lower(v)
    is_warn = k in SEVERITY_IS_WARNING_MARKERS
    if k in SEVERITY_ALIASES:
        mapped = SEVERITY_ALIASES[k]
        # medium-high solo válido para risk_level, no severity
        if mapped == "medium-high":
            return "high", True, is_warn
        return mapped, True, is_warn
    if k in CANON_SEVERITY:
        return k, True, is_warn
    return v.strip(), False, is_warn


def canon_risk_level(v: str) -> tuple[str, bool]:
    if not v.strip():
        return "", True
    k = norm_lower(v)
    if k in RISK_LEVEL_ALIASES:
        return RISK_LEVEL_ALIASES[k], True
    if k in CANON_RISK_LEVEL:
        return k, True
    # Erratas conocidas: si el valor es un ID (PR-*, GT-*, etc.) es
    # contaminación cross-column. Fallback conservador a "medium-high"
    # y queda flag en data_quality_issues.csv.
    for pfx in ("PR-", "TS-", "TST-", "GT-", "SRC-", "EX-"):
        if v.strip().startswith(pfx):
            return "medium-high", True
    return v.strip(), False


def canon_structure(v: str) -> tuple[str, bool]:
    if not v.strip():
        return "", True
    k = norm_lower(v)
    if k in STRUCTURE_ALIASES:
        return STRUCTURE_ALIASES[k], True
    if k in CANON_STRUCTURE:
        return k, True
    # Fallback: buscar substring (heurística corto plazo)
    for marker, target in STRUCTURE_SUBSTRING_ORDER:
        if marker in k:
            return target, True
    return v.strip(), False


def canon_action(v: str) -> tuple[str, bool]:
    if not v.strip():
        return "", True
    k = norm_lower(v).strip(" .")
    if k in ACTION_ALIASES:
        return ACTION_ALIASES[k], True
    if v.strip() in CANON_ACTION:
        return v.strip(), True
    # Heurística substring para prosa: aprobado como corto plazo.
    # Los gates con prosa larga (24 encontrados) se clasifican por
    # substring — la contaminación queda flag en el CSV editorial.
    if "manual" in k or "validación profesional" in k or "validacion profesional" in k:
        return "MANUAL_REVIEW", True
    if "bloquear" in k or "block" in k or "no seleccionar" in k or "no automatiz" in k:
        return "BLOCK", True
    if "regres" in k or "asistencia" in k or "reducir intensidad" in k:
        return "REGRESS", True
    if "reducir" in k or "reduce" in k or "modific" in k:
        return "ADJUST_VOLUME", True
    if "sustitu" in k or "usar jalón" in k or "elegir" in k and "asistid" in k:
        return "SUBSTITUTE", True
    if "evitar combin" in k or "un estímulo principal" in k:
        return "DEPRIORITIZE", True
    if "reorder" in k or "mover" in k:
        return "REORDER", True
    return v.strip(), False


def canon_status(v: str) -> tuple[str, bool]:
    """Heurística por substring — corto plazo."""
    if not v.strip():
        return "draft", True  # sin status → draft por defecto
    k = norm_lower(v)
    for marker, target in STATUS_HEURISTIC_ORDER:
        if marker in k:
            return target, True
    return v.strip(), False


def canon_equipment_tokens(v: str) -> tuple[list[str], bool]:
    """Devuelve lista de tokens canónicos + mapped_ok. El string se
    normaliza y se buscan substrings en EQUIPMENT_SUBSTRING_MAP (que
    está ordenado longest-first para evitar colisiones)."""
    if not v.strip():
        return [], True
    # Filtrar contaminación cross-column obvia
    if len(v) > 60 and ("." in v or ";" in v):
        return [], False

    vlow = norm_lower(v)
    tokens: list[str] = []
    matched_at_all = False

    # Aplicar substrings en orden: longest match wins vía prioridad de lista
    for pattern, canon_list in EQUIPMENT_SUBSTRING_MAP:
        if pattern in vlow:
            matched_at_all = True
            for c in canon_list:
                if c not in tokens:
                    tokens.append(c)
            # No hacemos break — el string puede tener múltiples equipos

    # Match directo con tokens canónicos (por si alguien ya escribió correcto)
    for tok in CANON_EQUIPMENT_TOKENS:
        if tok in vlow.split():
            if tok not in tokens:
                tokens.append(tok)
            matched_at_all = True

    if not tokens:
        return [], False
    return tokens, matched_at_all


# ===========================================================================
# PARSER DE DOSIS · rangos + notes_dosage
# ===========================================================================

# Regex para rangos numéricos con separadores em-dash / en-dash / hyphen.
_RANGE_RE = re.compile(
    r"(?P<lo>\d+(?:[.,]\d+)?)\s*[–—-]\s*(?P<hi>\d+(?:[.,]\d+)?)"
)
_SINGLE_RE = re.compile(r"^\s*(?P<n>\d+(?:[.,]\d+)?)\s*$")


def _num(s: str) -> float:
    return float(s.replace(",", "."))


def _extract_range(s: str) -> tuple[float | None, float | None]:
    """Extrae (min, max) numérico. Si es un solo número, min=max=n."""
    m = _RANGE_RE.search(s)
    if m:
        return _num(m.group("lo")), _num(m.group("hi"))
    m2 = _SINGLE_RE.match(s)
    if m2:
        n = _num(m2.group("n"))
        return n, n
    return None, None


def parse_dosage_field(
    raw: str, unit_hint: str = ""
) -> tuple[float | None, float | None, str]:
    """
    Parser genérico para un campo de dosis (work_interval, rest_interval,
    sets, reps, intensity, frequency).

    Devuelve (min, max, notes). notes preserva TODO el texto que no era
    parte del rango numérico principal — es donde caen los matices de
    seguridad como "buffer", "por rep", "según fases", "compatible con
    MaxHangs", "en última rep del último set".

    Casos:
      "3–15 s"                    → (3, 15, "s")
      "5–15 s por rep"            → (5, 15, "s por rep")
      "70–80%; buffer casi 0"     → (70, 80, "%; buffer casi 0")
      "Alta; buffer 1–5 s"        → (None, None, "Alta; buffer 1–5 s")
      "2–8"                       → (2, 8, "")
      "1"                         → (1, 1, "")
      "Según fases"               → (None, None, "Según fases")
      "No especificado en la fuente" → (None, None, "no especificado")
    """
    if not raw or not raw.strip():
        return None, None, ""
    s = raw.strip()
    low = s.lower()
    if "no especificad" in low:
        return None, None, "no especificado"

    m = _RANGE_RE.search(s)
    if m:
        lo, hi = _num(m.group("lo")), _num(m.group("hi"))
        # Notes = texto restante limpio
        notes = (s[:m.start()] + s[m.end():]).strip()
        # Reducir separadores dobles
        notes = re.sub(r"\s+", " ", notes).strip(" ;,-·")
        return lo, hi, notes

    m2 = _SINGLE_RE.match(s)
    if m2:
        n = _num(m2.group("n"))
        return n, n, ""

    # No hay número parseable — todo el texto es cualitativo, va a notes
    return None, None, s


# ===========================================================================
# CROSS-COLUMN CONTAMINATION DETECTOR
# ===========================================================================

def detect_contamination(sheet: str, row: dict, row_id: str) -> list[dict]:
    """Detecta patrones específicos de contaminación cross-column.
    Devuelve lista de dicts con detalle para el CSV `data_quality_issues.csv`.
    """
    issues = []
    for col, value in row.items():
        if not value:
            continue
        v = value.strip()
        vlow = v.lower()

        # risk_level con ID de protocolo/otra entidad
        if col == "risk_level":
            for pfx in ("PR-", "TS-", "TST-", "GT-", "SRC-", "EX-"):
                if v.startswith(pfx):
                    issues.append({
                        "sheet": sheet, "row_id": row_id, "column": col,
                        "value": v, "kind": "risk_level_has_id",
                    })

        # equipment con prosa (contiene punto final y es largo)
        if col == "equipment" and len(v) > 50 and ("." in v or ";" in v):
            issues.append({
                "sheet": sheet, "row_id": row_id, "column": col,
                "value": v[:120], "kind": "equipment_has_prose",
            })

        # level_min / level_max con environment
        if col in {"level_min", "level_max"}:
            env_markers = ("casa", "gimnasio", "roca")
            if any(m in vlow for m in env_markers):
                issues.append({
                    "sheet": sheet, "row_id": row_id, "column": col,
                    "value": v, "kind": "level_has_environment",
                })

        # action con prosa (punto final o muy largo)
        if col == "action":
            if v.endswith(".") or len(v) > 40:
                issues.append({
                    "sheet": sheet, "row_id": row_id, "column": col,
                    "value": v[:100], "kind": "action_has_prose",
                })

        # app_status con anatomía
        if col in {"app_status", "status"}:
            anat_markers = ("serrato", "trapecio", "deltoides", "manguito",
                            "rotador", "infraespinoso", "escápula")
            if any(m in vlow for m in anat_markers):
                issues.append({
                    "sheet": sheet, "row_id": row_id, "column": col,
                    "value": v, "kind": "status_has_anatomy",
                })

    return issues


# ===========================================================================
# APLANADORES POR HOJA
# ===========================================================================

FORCED_MANUAL_REVIEW_IDS = {
    # F2.4 findings editoriales (aprobado por Giuliana 2026-07-16):
    # dejarlas en manual_review hasta que Giuliana las revise en v3.0.
    "EX-FIN-030",  # Suspensión máxima full crimp (nombre autoproclamado máximo)
    "EX-FIN-040",  # Mono bloqueado en app (nombre autoproclamado)
    # curacion-dedos-v1 (Giuliana 2026-07-19): EX-FIN-057 "Deload de dedos"
    # tiene dosage_default = "GT-FIN-002" (ID de gate en campo de dosis).
    # A manual_review hasta que se defina el protocolo canónico de deload.
    "EX-FIN-057",
    # curacion calentamiento-recuperacion v1 (Giuliana 2026-07-20):
    # EX-WAR-017 "Recruitment hang submáximo" fuera del pool default de
    # calentamiento — es cargador de dedos, no calentamiento neutro.
    "EX-WAR-017",
}

# Override explícito de risk_level por ID (aplicado post-canonicalize).
# curacion calentamiento-recuperacion v1 (Giuliana 2026-07-20):
# el v3.0 los tenía como "medium" — al revés de lo que debe pasar. Estas
# 60 filas son lo MÁS disponible del catálogo · deben pasar el filtro
# maxRiskLevel=low que se aplica para condición baja / reconstrucción /
# novato. Sin este override, calentamiento + recuperación se bloquean
# para el grupo que MÁS los necesita.
RISK_LEVEL_OVERRIDES = {
    # Calentamiento · las 5 que cargan dedos o muro suben a low-medium
    "EX-WAR-016": "low-medium",
    "EX-WAR-019": "low-medium",
    "EX-WAR-020": "low-medium",
    "EX-WAR-028": "low-medium",
    "EX-WAR-029": "low-medium",
    # (EX-WAR-017 va a manual_review vía FORCED_MANUAL_REVIEW_IDS)
    # El resto de calentamiento (24 filas) y recuperación (30 filas) → low
    # se aplica por prefijo abajo con la regla del catch-all.
}

# Prefijos que van a `low` por default (post-canonicalize).
# Solo si no están en RISK_LEVEL_OVERRIDES ni en manual_review.
RISK_LEVEL_LOW_PREFIXES = ("EX-WAR-", "EX-REC-")

# curacion-bulk-11cat-v1 (Giuliana 2026-07-20) · REGLA A.
# Default risk_level por categoría cuando el original está contaminado
# (texto de dolor/parada en vez de nivel). Aplicado solo si el crudo
# NO está en CANON_RISK_LEVEL. La contaminación se mueve a stop_signals
# antes de reasignar (si stop_signals estaba vacío).
CATEGORY_RISK_DEFAULT = {
    "movilidad": "low",
    "mental": "low",
    "tecnica-escalada": "low-medium",
    "hombros-escapulas": "medium",
    "antebrazo-muneca-codo": "medium",
}

# curacion-bulk-11cat-v1 (Giuliana 2026-07-20) · REGLA B.
# 19 IDs cuyo status v3.0 tiene un patrón de movimiento en vez de un
# estado válido — forzar a `active` post-canon. La prosa contaminada
# no se recupera (los patrones anatómicos no aportan info nueva; están
# implícitos en category + objective).
STATUS_FORCE_ACTIVE_IDS = {
    # 17 hombros con patrón de movimiento en status
    "EX-SHO-008", "EX-SHO-010", "EX-SHO-011", "EX-SHO-012",
    "EX-SHO-018", "EX-SHO-023", "EX-SHO-024", "EX-SHO-025",
    "EX-SHO-026", "EX-SHO-028", "EX-SHO-029", "EX-SHO-030",
    "EX-SHO-031", "EX-SHO-032", "EX-SHO-033", "EX-SHO-035",
    "EX-SHO-036",
    # 2 en draft de movilidad simple
    "EX-MOB-018",  # Downward dog dinámico
    "EX-FLX-013",  # Estiramiento de pantorrilla
}

# Flag "hábito diario" · Giuliana 2026-07-20 (curación (B)).
# Estos EX de recuperación NO son ejercicios programables con sets × reps ·
# son rutinas de cuidado (sueño, hidratación, registro de dolor/fatiga,
# cuidado de piel). El motor de /sesion los excluye por default en
# restrict-pool.ts:0. Cuando exista /cuidado (feature futura) los va a
# consumir con su propia UI de checklists.
#
# LISTA FINAL confirmada por Giuliana 2026-07-20 · 17 IDs.
# Deltas vs provisional (+8): 015, 016, 019, 021, 022, 024, 027, 030.
# Los 13 físicos que quedan en pool /sesion: REC-001-010, 017, 025, 026.
# (Ver docs/rediseno/curacion/recuperacion-habits-v1.md para el mapping
# completo. Copy genérico de las 30 EX-REC-* pendiente de re-escritura
# editorial · no bloquea el mini-test.)
RECOVERY_HABIT_IDS = {
    "EX-REC-011",  # Registro de dolor post sesión
    "EX-REC-012",  # Registro de fatiga post sesión
    "EX-REC-013",  # Diario de sueño
    "EX-REC-014",  # Plan de descanso entre sesiones
    "EX-REC-015",  # Día de descarga
    "EX-REC-016",  # Semana de descarga
    "EX-REC-018",  # Hidratación post sesión
    "EX-REC-019",  # Comida post sesión
    "EX-REC-020",  # Cuidado de piel
    "EX-REC-021",  # Manejo de flapper
    "EX-REC-022",  # Lima suave de callos
    "EX-REC-023",  # Sueño como recuperación
    "EX-REC-024",  # Siesta corta
    "EX-REC-027",  # Chequeo de readiness al día siguiente
    "EX-REC-028",  # Retorno tras fatiga alta
    "EX-REC-029",  # Reducción de volumen por dolor
    "EX-REC-030",  # Cierre de sesión con aprendizaje
}


def flatten_exercise_row(
    row: dict, unmapped: dict, contamination: list, auto_fix_stats: dict
) -> dict:
    """v3.0 (46 cols) → v3.1 APP_ExerciseCatalog (21 cols)."""
    ex_id = row.get("exercise_id", "")

    # F4.0c · Auto-corrección de contamination patronada en hombros-escapulas.
    # Se aplica ANTES del detector de contamination para bajar el conteo real.
    # Los patrones auto-corregidos son claros (PR-* en risk_level, prosa en
    # equipment como oración con punto final, anatomía en app_status). Los
    # casos borderline quedan para revisión editorial de Giuliana.
    cat_raw = (row.get("category") or "").strip().lower()
    is_shoulder = "hombros" in cat_raw or "escápulas" in cat_raw or "escapulas" in cat_raw
    if is_shoulder:
        _autofix_shoulder_row(row, auto_fix_stats)

    # Detectar contaminación (post-autofix)
    for c in detect_contamination("Exercises", row, ex_id):
        contamination.append(c)

    # Canonicalizar campos clave
    cat, cat_ok = canon_category(row.get("category", ""))
    if not cat_ok:
        unmapped["category"][row.get("category", "")] += 1
    risk, risk_ok = canon_risk_level(row.get("risk_level", ""))
    if not risk_ok:
        unmapped["risk_level"][row.get("risk_level", "")] += 1
    status, status_ok = canon_status(row.get("app_status", ""))
    if not status_ok:
        unmapped["status"][row.get("app_status", "")] += 1
    equip_tokens, equip_ok = canon_equipment_tokens(row.get("equipment", ""))
    if not equip_ok and row.get("equipment", "").strip():
        unmapped["equipment"][row.get("equipment", "")] += 1
        # Regla F1.5 (Giuliana): equipment no mapeado → forzar
        # manual_review para que el motor no lo prescriba automáticamente.
        status = "manual_review"

    # F4.0b · Findings editoriales F2.4 → force manual_review por ID.
    if ex_id in FORCED_MANUAL_REVIEW_IDS:
        status = "manual_review"

    # curacion-bulk-11cat-v1 (Giuliana 2026-07-20) · REGLA B.
    # 19 IDs con patrón de movimiento en status → forzar active.
    # Corre DESPUÉS de FORCED_MANUAL_REVIEW · si un ID cae en ambos,
    # gana manual_review (ninguno se solapa hoy pero futura-proof).
    if ex_id in STATUS_FORCE_ACTIVE_IDS and status != "manual_review":
        status = "active"

    # curacion-bulk-11cat-v1 (Giuliana 2026-07-20) · REGLA A.
    # risk_level contaminado con prosa (dolor / parada) → mover a
    # stop_signals si estaba vacío, luego asignar default por categoría.
    # 72 filas afectadas: movilidad 33 · técnica 20 · hombros 12 ·
    # antebrazo 6 · mental 1. `risk_ok` es False cuando canon_risk_level
    # no matcheó → el `risk` var contiene el crudo (texto contaminado).
    if not risk_ok and cat_ok:
        raw_risk = row.get("risk_level", "").strip()
        if raw_risk and not row.get("stop_signals", "").strip():
            row["stop_signals"] = raw_risk  # promovido a stop_signals
        default_risk = CATEGORY_RISK_DEFAULT.get(cat)
        if default_risk:
            risk = default_risk
            # Quitar del contador de unmapped: ya lo resolvimos.
            unmapped["risk_level"][row.get("risk_level", "")] -= 1

    # curacion calentamiento-recuperacion v1 · risk_level overrides.
    # Solo aplica a activos (los manual_review ya salieron del pool).
    if status == "active":
        if ex_id in RISK_LEVEL_OVERRIDES:
            risk = RISK_LEVEL_OVERRIDES[ex_id]
        elif any(ex_id.startswith(p) for p in RISK_LEVEL_LOW_PREFIXES):
            risk = "low"

    return {
        "exercise_id": ex_id,
        "name_es": row.get("canonical_name_es", ""),
        "category": cat,
        # subcategory ELIMINADA (aprobado por Giuliana)
        "objective": row.get("objective", ""),
        # level_min / level_max reservados; v3.0 no los tiene explícitos
        "level_min": "",
        "level_max": "",
        "environment": row.get("subcategory", ""),  # placeholder heurístico
        # curacion-bulk-11cat-v1 · REGLA C2. Si tras canonicalizar no
        # queda ningún token en canon, emitir vacío (blanqueo). Antes se
        # preservaba el crudo (~130 tokens de prosa filtrados aquí).
        # `equipmentOk` en restrict-pool trata equip vacío como "pasa".
        "equipment": "; ".join(equip_tokens),
        "execution_summary": row.get("execution_summary", ""),
        "dosage_default": row.get("dosage_link", ""),  # se enlaza a protocolo
        "progression": row.get("progression", ""),
        "regression": row.get("regression", ""),
        "risk_level": risk,
        "stop_signals": row.get("stop_signals", ""),
        "gates": "",  # se puebla por JOIN post-regeneración (F1.3)
        "source_trace": row.get("source_id", ""),
        "source_sheet": "Exercises",
        "sprint": _sprint_from_row(row),
        "status": status,
        "validation_status": row.get("validation_note", ""),
        # Flag "hábito diario" (Giuliana 2026-07-20 · curación B).
        # Vacío para 553 EX · "habit" para los ~9 EX-REC-* de rutinas
        # de cuidado. El motor de /sesion excluye habits del pool.
        "type": "habit" if ex_id in RECOVERY_HABIT_IDS else "",
    }


def flatten_protocol_row(
    row: dict, unmapped: dict, contamination: list
) -> dict:
    """v3.0 Protocols (34 cols) → v3.1 APP_ProtocolCatalog (22 cols)
    con parser de dosis a rangos + notes_dosage."""
    pr_id = row.get("protocol_id", "")

    for c in detect_contamination("Protocols", row, pr_id):
        contamination.append(c)

    struct, struct_ok = canon_structure(row.get("session_position", ""))
    if not struct_ok:
        unmapped["structure"][row.get("session_position", "")] += 1
    risk, risk_ok = canon_risk_level(row.get("risk_level", ""))
    if not risk_ok:
        unmapped["risk_level"][row.get("risk_level", "")] += 1
    status, _ = canon_status(row.get("app_status", ""))
    # Regla F1.5 (Giuliana): structure no mapeado (residuales de
    # forearm/tracción) → manual_review. Se resuelve en Fase 4.
    if not struct_ok:
        status = "manual_review"

    # Parser de dosis · rangos
    work_min, work_max, work_notes = parse_dosage_field(row.get("work_interval", ""))
    rest_min, rest_max, rest_notes = parse_dosage_field(row.get("rest_interval", ""))
    sets_min, sets_max, sets_notes = parse_dosage_field(row.get("sets", ""))
    reps_min, reps_max, reps_notes = parse_dosage_field(row.get("reps", ""))
    int_min, int_max, int_notes = parse_dosage_field(row.get("intensity", ""))
    freq_min, freq_max, freq_notes = parse_dosage_field(row.get("frequency", ""))

    # Consolidar notas de dosis preservando matices de seguridad
    all_notes = "; ".join(filter(None, [
        f"work: {work_notes}" if work_notes else "",
        f"rest: {rest_notes}" if rest_notes else "",
        f"sets: {sets_notes}" if sets_notes else "",
        f"reps: {reps_notes}" if reps_notes else "",
        f"intensity: {int_notes}" if int_notes else "",
        f"frequency: {freq_notes}" if freq_notes else "",
    ]))

    # Promoción a manual_review si 0/6 campos numéricos poblados
    numeric_populated = sum(
        1 for v in [work_min, rest_min, sets_min, reps_min, int_min, freq_min]
        if v is not None
    )
    if numeric_populated == 0 and status == "active":
        status = "manual_review"

    return {
        "protocol_id": pr_id,
        "name_es": row.get("canonical_name_es", ""),
        "category": "",  # se deriva por link a ejercicios
        "objective": row.get("objective", ""),
        "eligible_level": row.get("eligible_level", ""),
        "exercise_ids": row.get("exercise_ids", ""),
        "structure": struct,
        "work": row.get("work_interval", ""),  # raw preservado
        "work_min_s": work_min,
        "work_max_s": work_max,
        "rest": row.get("rest_interval", ""),
        "rest_min_s": rest_min,
        "rest_max_s": rest_max,
        "sets": row.get("sets", ""),
        "sets_min": sets_min,
        "sets_max": sets_max,
        "reps": row.get("reps", ""),
        "reps_min": reps_min,
        "reps_max": reps_max,
        "intensity": row.get("intensity", ""),
        "intensity_pct_min": int_min,
        "intensity_pct_max": int_max,
        "notes_dosage": all_notes,
        "frequency": row.get("frequency", ""),
        "progression": row.get("progression_rule", ""),
        "regression": row.get("regression_rule", ""),
        "risk_level": risk,
        "gates": "",
        "source_trace": row.get("source_id", ""),
        "source_sheet": "Protocols",
        "sprint": _sprint_from_row(row),
        "status": status,
        "validation_status": row.get("notes", ""),
    }


def flatten_test_row(row: dict, unmapped: dict, contamination: list) -> dict:
    ts_id = row.get("test_id", "")
    for c in detect_contamination("Tests", row, ts_id):
        contamination.append(c)
    risk, risk_ok = canon_risk_level(row.get("risk_level", ""))
    if not risk_ok:
        unmapped["risk_level"][row.get("risk_level", "")] += 1
    status, _ = canon_status(row.get("app_status", ""))
    return {
        "test_id": ts_id,
        "name_es": row.get("canonical_name_es", ""),
        "category": "",
        "capacity": row.get("capacity", ""),
        "equipment": row.get("equipment", ""),
        "procedure_summary": row.get("protocol_summary", ""),
        "output_metric": row.get("output_metric", ""),
        "preconditions": row.get("preconditions", ""),
        "warmup": row.get("warmup", ""),
        "attempts": row.get("attempts", ""),
        "rest": row.get("rest", ""),
        "risk_level": risk,
        "gate_required": row.get("notes", ""),
        "source_trace": row.get("source_id", ""),
        "source_sheet": "Tests",
        "sprint": _sprint_from_row(row),
        "status": status,
        "validation_status": row.get("notes", ""),
    }


def flatten_gate_row(row: dict, unmapped: dict, contamination: list) -> dict:
    gt_id = row.get("gate_id", "")
    for c in detect_contamination("Gates", row, gt_id):
        contamination.append(c)
    action, action_ok = canon_action(row.get("action", ""))
    if not action_ok:
        unmapped["action"][row.get("action", "")] += 1
    severity, sev_ok, is_warn = canon_severity(row.get("severity", ""))
    if not sev_ok:
        unmapped["severity"][row.get("severity", "")] += 1
    status, _ = canon_status(row.get("status", ""))
    # Regla F1.5 (Giuliana): action no mapeado (residuales de
    # forearm/tracción) → manual_review. Se resuelve en Fase 4.
    if not action_ok:
        status = "manual_review"
    return {
        "gate_id": gt_id,
        "name": row.get("gate_name", ""),
        "category": row.get("basis_type", ""),
        "condition": row.get("condition_expression", ""),
        "action": action,
        "severity": severity,
        "is_warning": is_warn,
        "applies_to": row.get("applies_to", ""),
        "reason": row.get("user_message", ""),
        "requires_validation": (
            "yes" if "validación" in row.get("review_note", "").lower() else ""
        ),
        "source_trace": row.get("source_id", ""),
        "source_sheet": "Gates",
        "sprint": _sprint_from_row(row),
        "status": status,
        "validation_status": row.get("review_note", ""),
    }


def flatten_source_row(row: dict) -> dict:
    """v3.0 Sources → APP_SourceCatalog. Shape básico."""
    return {
        "source_id": row.get("source_id", ""),
        "title": row.get("title", ""),
        "author": row.get("author", ""),
        "year": row.get("year", ""),
        "type": row.get("type", ""),
        "use_in_app": row.get("use_in_app", ""),
        "confidence": row.get("confidence", ""),
        "file_or_origin": row.get("file_or_origin", ""),
        "url_or_doi": row.get("url_or_doi", ""),
        "notes": row.get("notes", ""),
        "status": canon_status(row.get("status", ""))[0],
    }


def _sprint_from_row(row: dict) -> str:
    """Heurística para inferir el sprint fuente. v3.0 no lo persiste
    por-fila; deducimos por prefijo de ID."""
    for id_col in ("exercise_id", "protocol_id", "test_id", "gate_id"):
        v = row.get(id_col, "")
        if v.startswith("EX-FIN") or v.startswith("PR-HB") or v.startswith("PR-FIN") or v.startswith("GT-FIN"):
            return "Sprint 6 (dedos/hangboard)"
        if v.startswith("EX-PULL") or v.startswith("PR-PULL") or v.startswith("GT-PULL"):
            return "Sprint 7 (traccion)"
        if v.startswith("EX-SHO") or v.startswith("EX-SCAP") or v.startswith("GT-SHO"):
            return "Sprint 8 (hombros)"
    return ""


# ===========================================================================
# AUTO-CORRECCIÓN CONTAMINATION HOMBROS-ESCAPULAS · F4.0c
# ===========================================================================

# Patrones anatómicos comunes que aparecen en `app_status` cuando deberían
# estar en `secondary_muscle_group` (que no existe en v3.1 · los movemos a
# `objective` como contexto conservado).
_ANATOMY_TOKENS = (
    "serrato", "trapecio", "deltoides", "manguito", "rotador",
    "infraespinoso", "escápula", "escapula", "romboides", "supraespinoso",
    "redondo menor", "dorsal", "pectoral",
)

# Prosa en `equipment` que en realidad es texto de expected_feeling o
# execution_summary. Se detecta por: (a) contiene punto final y >30 chars,
# o (b) contiene ";" con oración normal.
def _looks_like_prose(v: str) -> bool:
    if len(v) < 30:
        return False
    return v.rstrip().endswith(".") or (";" in v and len(v) > 40)


def _autofix_shoulder_row(row: dict, stats: dict) -> None:
    """Aplica reglas de auto-corrección a una fila de hombros-escapulas.
    Muta la fila en sitio; los movimientos quedan trazados en `stats`."""
    ex_id = row.get("exercise_id", "?")

    # Regla 1: PR-* en risk_level → moverlo a `dosage_link` (donde deberían
    # vivir las referencias a protocolos) + limpiar risk_level.
    risk_raw = row.get("risk_level", "").strip()
    if any(risk_raw.startswith(p) for p in ("PR-", "TS-", "TST-", "GT-", "SRC-", "EX-")):
        # El regenerador ya mapea PR-* → medium-high por fallback. Además,
        # persistimos la referencia a protocolo en dosage_link para no perder.
        existing_dose = row.get("dosage_link", "").strip()
        row["dosage_link"] = (
            f"{existing_dose}; {risk_raw}" if existing_dose else risk_raw
        )
        row["risk_level"] = "medium-high"  # fallback conservador
        stats["moved_pr_to_dosage_link"] = stats.get("moved_pr_to_dosage_link", 0) + 1
        stats.setdefault("moved_rows", []).append(f"{ex_id}: PR→dosage")

    # Regla 2: anatomía en app_status → moverla a `objective` (conservar
    # como contexto) + status crudo queda vacío → heurística lo pondrá en
    # 'active' o 'manual_review' según reglas normales.
    status_raw = row.get("app_status", "").strip()
    status_low = status_raw.lower()
    if any(tok in status_low for tok in _ANATOMY_TOKENS):
        obj_existing = row.get("objective", "").strip()
        # Solo conservar si el objective no ya menciona la anatomía
        if status_raw and status_raw not in obj_existing:
            row["objective"] = (
                f"{obj_existing} · {status_raw}" if obj_existing else status_raw
            )
        row["app_status"] = "Lista con restricciones"  # default active-ish
        stats["moved_anatomy_to_objective"] = stats.get("moved_anatomy_to_objective", 0) + 1
        stats.setdefault("moved_rows", []).append(f"{ex_id}: anatomy→objective")

    # Regla 3: prosa en equipment → moverla a expected_feeling (existente
    # o nueva) + equipment queda vacío → tokenizer canónico lo dejará como
    # empty (que pasa el gate como bodyweight).
    equip_raw = row.get("equipment", "").strip()
    if _looks_like_prose(equip_raw):
        feel = row.get("expected_feeling", "").strip()
        if equip_raw not in feel:
            row["expected_feeling"] = (
                f"{feel} · {equip_raw}" if feel else equip_raw
            )
        row["equipment"] = ""  # sin equipo canonicalizable
        stats["moved_prose_from_equipment"] = stats.get("moved_prose_from_equipment", 0) + 1
        stats.setdefault("moved_rows", []).append(f"{ex_id}: prose→feeling")


# ===========================================================================
# EXPANDER DE RELATIONSHIPS
# ===========================================================================

def expand_relationships(rows: list[dict]) -> list[dict]:
    """v3.0 Relationships (889 filas — headers: relationship_id,
    from_entity, from_id, relationship, to_entity, to_id, notes) →
    v3.1 APP_Relationships con columna `relation` renombrada + expand
    de listas en `to_id` si las hubiera."""
    out: list[dict] = []
    seq = 0
    for row in rows:
        from_id = row.get("from_id", "").strip()
        # v3.0 usa `relationship`; v3.1 se llama `relation`
        relation = (row.get("relationship") or row.get("relation") or "").strip()
        to_id_raw = row.get("to_id", "").strip()
        rel_id = row.get("relationship_id", "").strip()
        if not (from_id and relation and to_id_raw):
            continue
        # Splitear listas por `; ` o `|` (raro pero defensivo)
        to_ids = [t.strip() for t in re.split(r"[;|]", to_id_raw) if t.strip()]
        for i, tid in enumerate(to_ids):
            seq += 1
            # Preservar el rel_id de v3.0 cuando es una sola fila para
            # trazabilidad; si expandimos varias, agregamos sufijo.
            out_id = rel_id if len(to_ids) == 1 else f"{rel_id}-{i+1}"
            if not out_id:
                out_id = f"REL-{seq:06d}"
            out.append({
                "relationship_id": out_id,
                "from_id": from_id,
                "relation": relation,
                "to_id": tid,
                "category": row.get("from_entity", ""),
                "notes": row.get("notes", ""),
                "source_sheet": "Relationships",
                "origin": "v3.0 expand",
            })
    return out


# ===========================================================================
# GENERADORES DE HOJAS DERIVADAS
# ===========================================================================

def generate_onboarding_map(profile_schema_rows: list[dict]) -> list[dict]:
    """v3.0 ProfileSchema (35 cols estructurados) → APP_OnboardingMap
    (fila por campo del onboarding con used_by y related_gates)."""
    out: list[dict] = []
    for r in profile_schema_rows:
        field_id = r.get("field_id", "")
        if not field_id:
            continue
        out.append({
            "field_id": r.get("field_name", "").lower().replace(" ", "_"),
            "question_es": r.get("field_name", ""),
            "data_type": r.get("data_type", ""),
            "used_by": r.get("engine_use", ""),
            "related_gates": r.get("source_ref", ""),
            "notes": r.get("unknown_behavior", ""),
        })
    return out


def generate_session_builder_inputs(pipeline_rows: list[dict]) -> list[dict]:
    """v3.0 SelectionPipeline (12 stages) → APP_SessionBuilderInputs
    (inputs canónicos del motor con fail_closed_rule)."""
    out: list[dict] = []
    for r in pipeline_rows:
        stage = r.get("stage_id", "")
        if not stage:
            continue
        out.append({
            "input_id": r.get("stage_name", "").lower().replace(" ", "_"),
            "description": r.get("deterministic_operation", ""),
            "required": "Sí",
            "source_sheet": "SelectionPipeline",
            "used_for": r.get("output", ""),
            "fail_closed_rule": r.get("fail_closed_condition", ""),
        })
    return out


# ===========================================================================
# I/O · LOAD + WRITE
# ===========================================================================

def load_workbook_readonly(path: Path):
    return load_workbook(path, data_only=True, read_only=True)


def sheet_to_dicts(ws) -> list[dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return []
    headers = [str(h) if h is not None else "" for h in headers]
    # Detectar banner en primera fila (una sola celda, resto None) y buscar
    # el header real más abajo. Sprint13_Exercises usa este patrón.
    if len([h for h in headers if h and not h.startswith("Sprint")]) <= 1:
        for r in rows_iter:
            values = [c for c in r if c is not None]
            if len(values) >= 3:  # heurística: header real tiene ≥3 cols
                headers = [str(c) if c is not None else "" for c in r]
                break
    out: list[dict[str, str]] = []
    for r in rows_iter:
        if not any(c is not None for c in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            if i < len(r) and r[i] is not None:
                row[h] = str(r[i])
            else:
                row[h] = ""
        out.append(row)
    return out


# Mapeo de headers de Sprint*_Exercises → Exercises (main). Aprobado
# F4.0 (2026-07-16): algunas hojas sprint usan nombres ligeramente
# distintos que hay que normalizar antes del pipeline canónico.
SPRINT_EX_HEADER_MAP = {
    "canonical_name": "canonical_name_es",
    "execution": "execution_summary",
    "requires_validation": "app_status",  # se remapeará a status en canonicalización
    "risk_notes": "validation_note",
    "eligibility": "minimum_user_state",
    "gates_required": "prerequisite_ids",
}


def normalize_sprint_headers(rows: list[dict], mapping: dict[str, str]) -> list[dict]:
    """Renombra keys de una hoja sprint al vocabulario de main."""
    out = []
    for row in rows:
        new_row = {}
        for k, v in row.items():
            new_key = mapping.get(k, k)
            new_row[new_key] = v
        out.append(new_row)
    return out


def merge_all_exercises(wb) -> list[dict[str, str]]:
    """Fase 4 · lee Exercises + Sprint*_Exercises con dedupe por ID.
    Prefiere main (Sprint 1-12) sobre las hojas sprint sueltas."""
    main_rows = sheet_to_dicts(wb["Exercises"]) if "Exercises" in wb.sheetnames else []
    seen = {r["exercise_id"] for r in main_rows if r.get("exercise_id")}
    merged = list(main_rows)

    for sheet in wb.sheetnames:
        if not (sheet.startswith("Sprint") and sheet.endswith("_Exercises")):
            continue
        sprint_rows = sheet_to_dicts(wb[sheet])
        sprint_rows = normalize_sprint_headers(sprint_rows, SPRINT_EX_HEADER_MAP)
        added = 0
        for r in sprint_rows:
            ex_id = (r.get("exercise_id") or "").strip()
            if not ex_id or ex_id in seen:
                continue
            seen.add(ex_id)
            merged.append(r)
            added += 1
        if added:
            print(f"  merged {sheet}: +{added}")
    return merged


def merge_all_by_pattern(wb, main_name: str, sprint_pattern: str) -> list[dict]:
    """Genérico: main + Sprint*_<Pattern> con dedupe por primer campo ID."""
    main_rows = sheet_to_dicts(wb[main_name]) if main_name in wb.sheetnames else []
    if not main_rows:
        return []
    id_field = list(main_rows[0].keys())[0]
    seen = {r.get(id_field, "") for r in main_rows if r.get(id_field)}
    merged = list(main_rows)
    for sheet in wb.sheetnames:
        if not (sheet.startswith("Sprint") and sheet.endswith(f"_{sprint_pattern}")):
            continue
        for r in sheet_to_dicts(wb[sheet]):
            v = r.get(id_field, "").strip()
            if not v or v in seen:
                continue
            seen.add(v)
            merged.append(r)
    return merged


def write_v31_xlsx(target: Path, sheets: dict[str, list[dict]]) -> None:
    """Crea xlsx v3.1 con las hojas APP_*."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])

    wb.save(target)


# ===========================================================================
# ORQUESTADOR
# ===========================================================================

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path,
                        default=Path("docs/biblioteca-maestra-v3.0-consolidada.xlsx"))
    parser.add_argument("--target", type=Path,
                        default=Path("docs/biblioteca-maestra-v3.1.xlsx"))
    parser.add_argument("--outdir", type=Path,
                        default=Path("docs/rediseno/reports/f14_regenerate"))
    parser.add_argument("--skip-qa", action="store_true")
    args = parser.parse_args()

    if not args.source.exists():
        print(f"ERROR: fuente no encontrada: {args.source}", file=sys.stderr)
        return 2

    args.outdir.mkdir(parents=True, exist_ok=True)

    print(f"Cargando {args.source}...")
    wb = load_workbook_readonly(args.source)

    def load(name: str) -> list[dict[str, str]]:
        if name not in wb.sheetnames:
            return []
        return sheet_to_dicts(wb[name])

    # F4.0 · merge main + Sprint*_ para las 5 entidades principales.
    print("Merge de main + Sprint*_*:")
    exercises_v30 = merge_all_exercises(wb)
    protocols_v30 = merge_all_by_pattern(wb, "Protocols", "Protocols")
    tests_v30 = merge_all_by_pattern(wb, "Tests", "Tests")
    gates_v30 = merge_all_by_pattern(wb, "Gates", "Gates")
    sources_v30 = merge_all_by_pattern(wb, "Sources", "Sources")
    relationships_v30 = merge_all_by_pattern(wb, "Relationships", "Relations")
    profile_schema_v30 = load("ProfileSchema")
    pipeline_v30 = load("SelectionPipeline")

    print(f"\nTotales post-merge:")
    print(f"  {len(exercises_v30)} EX · {len(protocols_v30)} PR · "
          f"{len(tests_v30)} TS · {len(gates_v30)} GT · "
          f"{len(sources_v30)} SRC · {len(relationships_v30)} REL")

    # Contadores de unmapped values
    unmapped: dict[str, Counter] = defaultdict(Counter)
    contamination: list[dict] = []
    auto_fix_stats: dict = {}

    print("\nAplanando + canonicalizando...")
    exercises_v31 = [
        flatten_exercise_row(r, unmapped, contamination, auto_fix_stats)
        for r in exercises_v30
    ]
    protocols_v31 = [
        flatten_protocol_row(r, unmapped, contamination) for r in protocols_v30
    ]
    tests_v31 = [
        flatten_test_row(r, unmapped, contamination) for r in tests_v30
    ]
    gates_v31 = [
        flatten_gate_row(r, unmapped, contamination) for r in gates_v30
    ]
    sources_v31 = [flatten_source_row(r) for r in sources_v30]

    print("Expandiendo Relationships...")
    relationships_v31 = expand_relationships(relationships_v30)
    print(f"  {len(relationships_v30)} → {len(relationships_v31)} filas")

    print("Generando hojas derivadas...")
    onboarding_map_v31 = generate_onboarding_map(profile_schema_v30)
    session_builder_v31 = generate_session_builder_inputs(pipeline_v30)

    print(f"\nEscribiendo {args.target}...")
    sheets = {
        "APP_ExerciseCatalog": exercises_v31,
        "APP_ProtocolCatalog": protocols_v31,
        "APP_TestCatalog": tests_v31,
        "APP_GateCatalog": gates_v31,
        "APP_SourceCatalog": sources_v31,
        "APP_Relationships": relationships_v31,
        "APP_OnboardingMap": onboarding_map_v31,
        "APP_SessionBuilderInputs": session_builder_v31,
    }
    write_v31_xlsx(args.target, sheets)

    # Reporte de unmapped values (ampliar mapping tables en próximo sprint)
    unmapped_csv = args.outdir / "unmapped_values.csv"
    with unmapped_csv.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["vocabulary", "raw_value", "count"])
        for vocab, counter in unmapped.items():
            for v, n in counter.most_common():
                w.writerow([vocab, v, n])
    print(f"  Unmapped values: {unmapped_csv}")

    # CSV de contaminación cross-column (44 filas esperadas)
    contamination_csv = args.outdir / "data_quality_issues.csv"
    with contamination_csv.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row_id", "column", "value", "kind"])
        for i in contamination:
            w.writerow([i["sheet"], i["row_id"], i["column"],
                       i["value"], i["kind"]])
    print(f"  Contamination: {contamination_csv} ({len(contamination)} filas)")

    # Métricas resumen
    total_unmapped = sum(sum(c.values()) for c in unmapped.values())
    print(f"\nMétricas:")
    print(f"  Sheets escritas: {len(sheets)}")
    print(f"  Contamination detectada (post-autofix): {len(contamination)} filas")
    print(f"  Auto-fix hombros-escapulas:")
    for k, v in auto_fix_stats.items():
        if k == "moved_rows":
            continue
        print(f"    {k}: {v}")
    print(f"  Unmapped values: {total_unmapped}")
    for vocab, counter in unmapped.items():
        print(f"    {vocab}: {sum(counter.values())} ({len(counter)} valores únicos)")

    # QA post-regeneración
    if not args.skip_qa:
        print("\nCorriendo QA sobre v3.1 regenerada...")
        import subprocess
        result = subprocess.run(
            [sys.executable, "scripts/rediseno/qa_audit_v30.py",
             "--source", str(args.target),
             "--outdir", "docs/rediseno/reports/f14_regenerate_qa"],
            capture_output=True, text=True,
        )
        print(result.stdout[-500:])
        if result.returncode != 0:
            print(result.stderr[-500:], file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
