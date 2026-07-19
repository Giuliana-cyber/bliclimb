#!/usr/bin/env python3
"""
xlsx → JSON bundler para el motor invertido.

Lee `docs/biblioteca-maestra-v3.1.xlsx` (output de regenerate_v31.py)
y produce `data/catalog-v3.1.json` que consume `lib/brain/motor-inverted/
catalog-loader.ts`.

También lee `data/focus-rules.json` si existe y lo inyecta como
`FocusRules` en el JSON de salida (Fase 3 · F3.1).

Uso:
    python3 scripts/rediseno/xlsx_to_json.py
    python3 scripts/rediseno/xlsx_to_json.py --xlsx <path> --out <path>
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl


def sheet_to_dicts(ws) -> list[dict]:
    """Convierte una hoja en lista de dicts. Encabezados en fila 1.
    Todo valor → str vacío si es None (el TS coerce con `|| undefined`)."""
    headers = [c.value for c in ws[1]]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None or c == "" for c in row):
            continue
        rows.append({h: ("" if v is None else str(v)) for h, v in zip(headers, row) if h})
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="docs/biblioteca-maestra-v3.1.xlsx")
    ap.add_argument("--out", default="data/catalog-v3.1.json")
    ap.add_argument("--focus-rules", default="data/focus-rules.json")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    if not xlsx_path.exists():
        print(f"[!] xlsx no encontrado: {xlsx_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    payload: dict = {}
    # Todas las hojas APP_* van tal cual
    for sh_name in wb.sheetnames:
        if not sh_name.startswith("APP_"):
            continue
        payload[sh_name] = sheet_to_dicts(wb[sh_name])
        print(f"  {sh_name}: {len(payload[sh_name])} filas")

    # FocusRules · inyectar del JSON separado si existe
    fr_path = Path(args.focus_rules)
    if fr_path.exists():
        with fr_path.open() as f:
            payload["FocusRules"] = json.load(f)
        print(f"  FocusRules (inyectado): {len(payload['FocusRules'])}")
    else:
        print(f"  [!] {fr_path} no existe · FocusRules quedará vacío")
        payload["FocusRules"] = []

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    size_kb = out_path.stat().st_size // 1024
    print(f"\n[OK] {out_path} escrito ({size_kb} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
